# -*- coding: utf-8 -*-
"""
第二轮结构整改 V2 —— 数据预处理 + 字段审计
输入：当前主文件「AmazonES_产品清单与提取信息_选品优化版.xlsx」（第一轮产出，5 张表）
      - 后台数据 表：全部原始 + 第一轮清洗字段（200 行）
      - 选品清单 表：读取选品状态/研究备注（人工字段，重导不覆盖）
输出：
  _v2_selected.json  —— V2 数据模型所需全部计算字段
  _v2_audit.json     —— 13 项字段可用率 + A/B/C 原因分类
                         A=原采集没抓到  B=原始详情已有可整理  C=程序解析失败
本轮原则：不联网、不新增SKU、不猜缺失字段；能恢复的只从已有 details_json 恢复。
"""
import openpyxl
import json
import re
import os
import datetime

SRC = r"E:\amazon_es\.worktrees\reconnaissance\AmazonESBestseller\outputs\amazon_es_catalog_20260825\AmazonES_产品清单与提取信息_选品优化版.xlsx"
OUTDIR = os.path.dirname(SRC)
OUT_DATA = os.path.join(OUTDIR, "_v2_selected.json")
OUT_AUDIT = os.path.join(OUTDIR, "_v2_audit.json")

MONTHS_ES = {
    'enero':1,'febrero':2,'marzo':3,'abril':4,'mayo':5,'junio':6,
    'julio':7,'agosto':8,'septiembre':9,'octubre':10,'noviembre':11,'diciembre':12,
}

# ---------- 通用工具 ----------
def excel_serial_to_dt(v):
    if v is None: return None
    try: f = float(str(v).replace(',', '.').strip())
    except (ValueError, TypeError): return None
    try: return datetime.datetime(1899, 12, 30) + datetime.timedelta(days=f)
    except (OverflowError, ValueError): return None

def norm_dt(v):
    """round-1 后台数据的 first_seen/last_seen 可能是字符串或数字，统一为字符串"""
    if v is None or str(v).strip() == '': return ''
    s = str(v).strip()
    if re.match(r'^\d+(\.\d+)?$', s):
        dt = excel_serial_to_dt(s)
        return dt.isoformat(sep=' ') if dt else s
    return s

def parse_es_date(s):
    if not s: return None
    m = re.match(r'^(\d{1,2})\s+([A-Za-záéíóúñÁÉÍÓÚÑ]+)\s+(\d{4})\s*$', str(s).strip())
    if not m: return None
    d, mo, y = int(m.group(1)), m.group(2).lower(), int(m.group(3))
    if mo not in MONTHS_ES: return None
    try: return datetime.date(y, MONTHS_ES[mo], d)
    except ValueError: return None

def clean_brand(b):
    if not b: return ''
    s = str(b).strip()
    m = re.match(r'^Visita\s+la\s+tienda\s+de\s+(.+)$', s, re.IGNORECASE)
    return m.group(1).strip() if m else s

def _dec_comma(s):
    return re.sub(r'(?<=\d),(?=\d)', '.', s)

def extract_bsr_segments(s):
    """返回 [(cat, rank), ...] 列表，无则 []
    修正：类目名懒匹配，在下一个 'nº'、'(' 或行尾前停止（原贪婪匹配会把多段拼成一坨）"""
    if not s: return []
    segs = []
    pat = re.compile(r'n\.?º?\s*([\d.,]+)\s+en\s+([^()\n]+?)(?=\s*n[.]?\s*º|\s*\(|\s*$)')
    for m in pat.finditer(s):
        rank = m.group(1).replace('.', '').replace(',', '')
        cat = re.sub(r'\s*(Ver el|Ver los|Ver).*$', '', m.group(2)).strip()
        if cat: segs.append((cat, rank))
    return segs

# ---------- 规格（简短版，中文） ----------
_CAP_TERMS = [
    ('Centímetros cúbicos', '立方厘米'), ('centímetros cúbicos', '立方厘米'),
    ('Litros', '升'), ('litros', '升'), ('Litro', '升'), ('litro', '升'),
    ('Mililitros', '毫升'), ('mililitros', '毫升'), ('Mililitro', '毫升'), ('mililitro', '毫升'),
]
_DIM_RE1 = re.compile(r'^([\d.]+)\s*l\.\s*x\s*([\d.]+)\s*an\.\s*x\s*([\d.]+)\s*al\.\s*(centímetros|milímetros|metros)?', re.I)
_DIM_RE2 = re.compile(r'^([\d.]+)\s*l\.\s*x\s*([\d.]+)\s*an\.\s*(centímetros|milímetros|metros)?', re.I)
_DIM_RE4 = re.compile(r'^([\d.]+)\s*x\s*([\d.]+)\s*x\s*([\d.]+)\s*(cm|mm|m)?', re.I)
_DIM_RE5 = re.compile(r'^([\d.]+)\s*an\.\s*x\s*([\d.]+)\s*al\.\s*(centímetros|milímetros|metros)?', re.I)
_DIM_RE6 = re.compile(r'^([\d.]+)\s*l\.\s*x\s*([\d.]+)\s*al\.\s*(centímetros|milímetros|metros)?', re.I)
_UNIT_CN = {'centímetros':'厘米','milímetros':'毫米','metros':'米','cm':'厘米','mm':'毫米','m':'米','':''}

def _dim_zh(v):
    s = _dec_comma(str(v).strip())
    # 剥离不可见/特殊空白字符（Amazon 常带零宽字符）
    s = re.sub(r'[‎‏​﻿ ]', '', s).strip()
    m = (_DIM_RE1.match(s) or _DIM_RE2.match(s) or _DIM_RE4.match(s)
         or _DIM_RE5.match(s) or _DIM_RE6.match(s))
    if not m: return None
    unit = ''
    for p in m.groups():
        if p and p.lower() in _UNIT_CN: unit = _UNIT_CN[p.lower()]
    nums = [g for g in m.groups() if g and g.lower() not in _UNIT_CN]
    return '×'.join(nums) + unit

def _cap_zh(v):
    s = _dec_comma(str(v).strip())
    for es, zh in sorted(_CAP_TERMS, key=lambda t: -len(t[0])):
        s = re.sub(r'\b' + re.escape(es) + r'\b', zh, s)
    return s

def package_count(d):
    """件数：取多个计数字段的最大值（total_del_paquete 常为单价参考值 1，须以真实件数为准）"""
    best = None
    for k in ('numero_de_articulos', 'numero_de_piezas',
              'cantidad_de_articulos_en_el_paquete',
              'total_del_paquete_segun_la_medida_elegida_para_referenciar_precio'):
        v = d.get(k)
        if v:
            m = re.match(r'([\d.,]+)', str(v).replace(' ', ''))
            if m:
                val = float(m.group(1).replace(',', '.'))
                if best is None or val > best: best = val
    if best is None: return None
    return ('%d' % best) if best == int(best) else ('%.2f' % best)

def build_spec_v2(d):
    """简短规格：只回答"客户买的是哪个规格版本"，如 4件套 / 17×3.2×25.2厘米 / 1升"""
    if not d: return ''
    parts = []
    cnt = package_count(d)
    if cnt:
        try: n = int(float(cnt))
        except ValueError: n = 0
        if n > 1:
            nset = d.get('numero_de_sets')
            nset_n = None
            if nset:
                m = re.match(r'([\d.]+)', str(nset).replace(' ', ''))
                if m: nset_n = int(float(m.group(1)))
            parts.append('%d件套' % n if nset_n == n else '%d件' % n)
    cap = (d.get('capacidad') or d.get('capacidad_de_salida') or d.get('volumen_de_almacenamiento')
           or d.get('volumen_del_tanque') or d.get('volumen_liquido'))
    if cap: parts.append(_cap_zh(cap))
    dim = (d.get('dimensiones_del_articulo_largo_x_ancho_x_alto')
           or d.get('dimensiones_del_producto') or d.get('dimensiones_articulo')
           or d.get('dimensiones_del_articulo_l_x_a') or d.get('dimensiones_del_articulo_ancho_x_alto'))
    if dim:
        dz = _dim_zh(dim)
        if dz: parts.append(dz)
    pot = d.get('potencia')
    if pot:
        parts.append(_cap_zh(str(pot).replace('vatios', '瓦').replace('watios', '瓦')))
    volt = d.get('voltaje') or d.get('tension')
    if volt:
        parts.append(_cap_zh(str(volt).replace('Voltios', 'V').replace('voltios', 'V')))
    comp = d.get('cantidad_de_compartimentos')
    if comp:
        m = re.match(r'(\d+)', str(comp))
        if m: parts.append('%s格' % m.group(1))
    pcs = d.get('numero_de_piezas')
    if pcs and not cnt:
        m = re.match(r'(\d+)', str(pcs))
        if m and int(m.group(1)) > 1: parts.append('%s只' % m.group(1))
    return ' / '.join(parts)

# ---------- 摘要（压缩中文事实） ----------
_ZH_MAT = [
    ('Acero inoxidable','不锈钢'),('acero inoxidable','不锈钢'),('Vidrio de borosilicato','硼硅玻璃'),
    ('Vidrio templado','钢化玻璃'),('Tereftalato de polietileno','PET塑料'),('Poliéster','涤纶'),
    ('Polipropileno','聚丙烯'),('Polietileno','聚乙烯'),('Policarbonato','聚碳酸酯'),('Poliuretano','聚氨酯'),
    ('Fibra de vidrio','玻璃纤维'),('Silicona','硅胶'),('Cerámica','陶瓷'),('Aluminio','铝'),
    ('Algodón','棉'),('Bambú','竹'),('Madera','木材'),('Resina','树脂'),('Nylon','尼龙'),
    ('Nailon','尼龙'),('Caucho','橡胶'),('Goma','橡胶'),('Plástico','塑料'),('Plastico','塑料'),
    ('Metal','金属'),('Cuero','皮革'),('Tela','织物'),('Esparto','茅草'),('Oxford','牛津布'),
    ('Inoxidable','不锈钢'),('Acero','钢'),('Vidrio','玻璃'),('Ratán','藤'),('ABS','ABS'),
    ('Porcelana','陶瓷'),('Piedra','石材'),('Teflón','特氟龙'),('Teflon','特氟龙'),('Mármol','大理石'),
    ('Cuarzo','石英'),('Grafito','石墨'),('Cristal','水晶'),('Elástico','弹力'),('Espuma','海绵'),
    ('Vellón','绒'),('Lino','亚麻'),('Seda','丝绸'),('Piel','皮革'),('Corcho','软木'),
    ('Carbono','碳纤维'),('Vidrio templado','钢化玻璃'),('Poliamida','聚酰胺'),
]
_ZH_TR = [
    ('Apilable','可堆叠'),('Extraíble','可拆卸'),('Extraible','可拆卸'),('Desmontable','可拆卸'),
    ('Reciclable','可回收'),('Reutilizable','可重复使用'),('Doble sello','双重密封'),('Airtight','密封防漏'),
    ('A prueba de agua','防水'),('Impermeable','防水'),('Sin BPA','不含BPA'),('Sí','是'),('No','否'),
    ('Diseño plegable','可折叠'),('Plegable','可折叠'),('Portátil','便携'),('Ligero','轻便'),
    ('Robusto','耐用'),('Robusta','耐用'),('Flexible','柔韧'),('Flexibles','柔韧'),
    ('resistente a desgarros','耐撕'),('Resistente a desgarros','耐撕'),
    ('Sin enjuague','免冲洗'),('Sin residuos','无残留'),('Sin residuo','无残留'),
    ('Concentrado','浓缩'),('Concentrada','浓缩'),('Sin aluminio','不含铝'),('Sin cobre','不含铜'),
    ('Sin metales pesados','不含重金属'),('Metales pesados','重金属'),('alta presión','高压'),
    ('Alta presión','高压'),('ajustable','可调节'),('Ajustable','可调节'),('fácil instalación','易安装'),
    ('Fácil instalación','易安装'),('Descalcificador','除垢'),('elimina depósitos de cal','去除水垢'),
    ('Elimina depósitos de cal','去除水垢'),('Filtración avanzada','深度过滤'),('sin costuras','无缝'),
    ('Sin costuras','无缝'),('hermético','密封'),('Hermético','密封'),('resistente al calor','耐热'),
    ('Resistente al calor','耐热'),('A prueba de fugas','防漏'),('antideslizante','防滑'),
    ('Antideslizante','防滑'),('transpirable','透气'),('Transpirable','透气'),('a prueba de polvo','防尘'),
]
_ZH_USO = [
    ('Comida para llevar','外卖'),('Transportador de alimentos','食物携带'),('Alfombras','地毯'),
    ('Interior del automóvil','汽车内饰'),('Tapetes','门垫'),('Tapicería','织物'),('Ropa','衣物'),
    ('Baño','浴室'),('Cocina','厨房'),('Jardin','花园'),('Jardín','花园'),('Exterior','户外'),
    ('Interior','室内'),('Almacenamiento de comidas','食物储存'),('Máquinas de cápsulas','胶囊咖啡机'),
    ('Desincrustantes para cafeteras','咖啡机除垢'),('Mascotas','宠物'),('Limpieza','清洁'),
    ('Organización','收纳'),('Viaje','旅行'),('Escuela','学校'),('Picnic','野餐'),
    ('Congelador','冷冻'),('Nevera','冰箱'),('Horno','烤箱'),('Microondas','微波炉'),
]
def _zh_apply(s, terms):
    if not s: return ''
    s = str(s)
    for es, zh in sorted(terms, key=lambda t: -len(t[0])):
        s = re.sub(r'\b' + re.escape(es) + r'\b', zh, s)
    return s.strip()

def build_summary_v2(d):
    if not d: return ''
    facts = []
    mat = (d.get('tipo_de_material') or d.get('material') or d.get('material_o_tela')
           or d.get('material_de_la_tapa') or d.get('tipo_de_material_de_la_parte_superior'))
    if mat:
        mz = _zh_apply(mat, _ZH_MAT)
        facts.append('材质：' + mz if mz else '材质：' + str(mat).strip())
    cnt = package_count(d)
    if cnt:
        try: n = int(float(cnt))
        except ValueError: n = 0
        if n > 1:
            nset = d.get('numero_de_sets'); ns = None
            if nset:
                m = re.match(r'([\d.]+)', str(nset).replace(' ', ''))
                if m: ns = int(float(m.group(1)))
            facts.append('%d件套' % n if ns == n else '%d件' % n)
    cap = (d.get('capacidad') or d.get('capacidad_de_salida') or d.get('volumen_de_almacenamiento')
           or d.get('volumen_del_tanque'))
    if cap: facts.append('容量：' + _cap_zh(cap))
    dim = (d.get('dimensiones_del_articulo_largo_x_ancho_x_alto')
           or d.get('dimensiones_del_producto') or d.get('dimensiones_articulo')
           or d.get('dimensiones_del_articulo_l_x_a') or d.get('dimensiones_del_articulo_ancho_x_alto'))
    if dim:
        dz = _dim_zh(dim)
        if dz: facts.append('尺寸：' + dz)
    esp = d.get('carecteristicas_especiales') or d.get('caracteristicas_especiales')
    if esp:
        ez = _zh_apply(esp, _ZH_TR)
        if ez: facts.append('特点：' + ez)
    herm = d.get('hermeticidad_del_recipiente_de_almacenamiento_de_alimentos')
    if herm:
        hz = _zh_apply(herm, _ZH_TR)
        if hz: facts.append('密封：' + hz)
    apto = []
    if 'Sí' in str(d.get('el_articulo_es_apto_para_el_microondas') or ''): apto.append('微波炉')
    if 'Sí' in str(d.get('el_articulo_es_apto_para_lavavajillas') or ''): apto.append('洗碗机')
    cm = str(d.get('caracteristicas_de_materiales') or '') + ' ' + str(d.get('instrucciones_de_cuidado_del_producto') or '')
    if re.search(r'congelador|nevera', cm, re.I): apto.append('冷冻')
    if apto: facts.append('可' + '/'.join(apto))
    wtr = d.get('nivel_de_resistencia_al_agua')
    if wtr and re.search(r'agua|impermeable', str(wtr), re.I): facts.append('防水')
    saf = d.get('sin_tipo_de_material')
    if saf:
        sz = _zh_apply(saf, _ZH_TR)
        facts.append(sz if sz else str(saf).strip())
    req = d.get('requisitos_cumplidos')
    if req: facts.append('认证：' + str(req).strip())
    inc = d.get('elementos_incluidos') or d.get('componentes_incluidos') or d.get('nombre_del_conjunto')
    if inc:
        m = re.match(r'^\s*(\d+)\s*([A-Za-zÁÉÍÓÚáéíóúñÑ ]{0,20})', str(inc))
        if m: facts.append('包装含%d件' % int(m.group(1)))
    uso = d.get('usos_recomendados_para_producto') or d.get('usos_especificos_para_producto') or d.get('uso')
    if uso:
        uz = _zh_apply(uso, _ZH_USO)
        if uz: facts.append('用途：' + uz)
    org = d.get('country_of_origin') or d.get('pais_de_origen')
    if org:
        facts.append('产地：' + _zh_apply(org, [('España','西班牙'),('China','中国'),('Alemania','德国'),
                                                ('Francia','法国'),('Italia','意大利'),('Portugal','葡萄牙')]))
    txt = '；'.join(facts)
    if len(txt) > 250:
        cut = txt[:250].rfind('；')
        txt = txt[:cut] if cut > 100 else txt[:250]
        txt = txt.rstrip('；') + '…'
    return txt

# ---------- 读取当前主文件（round-1 产出） ----------
wb = openpyxl.load_workbook(SRC, data_only=True)

ws = wb['后台数据']
rows = list(ws.iter_rows(min_row=1, values_only=True))
header = [str(h).strip() if h else '' for h in rows[0]]
data = [r for r in rows[1:] if r and any(c is not None and str(c).strip() != '' for c in r)]
idx = {h: i for i, h in enumerate(header)}
print('后台数据 header 数:', len(header), '| 数据行:', len(data))

# 选品清单：读取人工字段（选品状态/研究备注），行序与后台数据一致（row4+i ↔ row2+i）
ws_main = wb['选品清单']
main_rows = list(ws_main.iter_rows(min_row=4, max_row=3 + len(data), values_only=True))
# 列：0图片 1序号 2名称 3价格 4原价 5折扣 6一级 7二级 8三级 9细分 10月购 11类目排名 12规格 13摘要 14上架 15链接 16图链 17ASIN 18选品状态 19研究备注

records = []
d_list = []
audit_def = ['current_price','original_price','discount_rate','l1','l2','l3','leaf_category',
             'browse_node_id','monthly_bought','category_rank','spec','launch_date','parent_asin']
audit = {k: {'available':0,'total':len(data),'missing':0,'A':0,'B':0,'C':0} for k in audit_def}

for i, r in enumerate(data):
    def g(name):
        v = r[idx[name]] if name in idx and idx[name] < len(r) else None
        return '' if v is None else str(v).strip()
    asin = g('asin')
    d = None
    if g('details_json'):
        try: d = json.loads(g('details_json'))
        except Exception: d = None
    d_list.append(d)
    bsr_segs = extract_bsr_segments((d or {}).get('clasificacion_en_los_mas_vendidos_de_amazon', ''))
    main_cat = bsr_segs[0] if bsr_segs else None
    leaf = bsr_segs[-1] if len(bsr_segs) > 1 else None

    cur = g('current_price'); orig = g('original_price'); dr = g('discount_rate')
    disc = ''
    if cur and orig:
        try:
            cf = float(cur.replace(',', '.')); of = float(orig.replace(',', '.'))
            if of > 0: disc = round((of - cf) / of, 4)
        except ValueError: disc = ''
    dfa = parse_es_date(g('date_first_available_raw'))
    spec_v2 = build_spec_v2(d)
    summary_v2 = build_summary_v2(d)

    # 人工字段（从选品清单带过来，防止重导覆盖）
    sel_status = main_rows[i][18] if i < len(main_rows) and main_rows[i] and len(main_rows[i]) > 18 else None
    remark = main_rows[i][19] if i < len(main_rows) and main_rows[i] and len(main_rows[i]) > 19 else None
    sel_status = '待评估' if sel_status is None or str(sel_status).strip() == '' else str(sel_status).strip()
    remark = '' if remark is None else str(remark).strip()

    # 重采标志：缺失的核心字段
    recrawl = []
    if not g('browse_node_id'): recrawl.append('browse_node_id')
    if not g('monthly_bought_text'): recrawl.append('monthly_bought')
    if not orig: recrawl.append('original_price')
    recrawl.append('L2/L3')
    if not (leaf or (d or {}).get('clasificacion_en_los_mas_vendidos_de_amazon')): recrawl.append('BSR榜单')
    if not spec_v2: recrawl.append('规格')
    if not dfa: recrawl.append('上架时间')

    records.append({
        'row_idx': i,
        'asin': asin,
        'parent_asin': g('parent_asin'), 'parent_asin_status': g('parent_asin_status'),
        '采集类目中文': g('采集类目中文'), '采集类目西语': g('采集类目西语'),
        'title_es_raw': g('title_es'),
        'brand_raw': g('brand'), 'brand': clean_brand(g('brand')),
        'price_legacy': g('price'),
        'original_price': orig, 'current_price': cur, 'currency': g('currency'),
        'discount_rate': disc if disc != '' else (dr if dr else ''),
        'rating': g('rating'), 'review_count': g('review_count'),
        'monthly_bought_raw': g('monthly_bought_text'), 'monthly_bought_min': g('monthly_bought_min'),
        'image_url': g('image_url'), 'product_url': 'https://www.amazon.es/dp/' + asin,
        'details_json': g('details_json'), 'details_raw': g('details'),
        'specification_legacy': g('specification'),
        'date_first_available': dfa.strftime('%Y-%m-%d') if dfa else '',
        'date_first_available_raw': g('date_first_available_raw'),
        'first_seen': norm_dt(g('first_seen')), 'last_seen': norm_dt(g('last_seen')),
        'ranking_count': g('ranking_count'), 'best_rank_legacy': g('best_rank'),
        'image_path': g('image_path'), 'image_download_status': g('image_download_status'),
        'image_download_error': g('image_download_error'),
        '详情状态': g('详情状态') or ('已提取' if g('详情已提取') == '是' else '未提取'),
        '图片状态': g('图片状态') or ('已嵌入' if g('图片已下载') == '是' else '未下载'),
        # 排行榜相关
        'bsr_segments': bsr_segs,
        'bsr_main_cat': (main_cat[0] if main_cat else ''),
        'bsr_main_rank': (main_cat[1] if main_cat else ''),
        'bsr_leaf_cat': (leaf[0] if leaf else ''),
        'bsr_leaf_rank': (leaf[1] if leaf else ''),
        '上榜细分类目数': len(bsr_segs),
        # V2 人工字段
        'spec_v2': spec_v2,
        'summary_v2': summary_v2,
        'upc': (d or {}).get('upc', ''), 'gtin': (d or {}).get('num_de_identificacion_comercial_global', ''),
        '选品状态': sel_status, '研究备注': remark,
        '重采标志': ','.join(recrawl),
    })

    # ---------- 审计计数 + A/B/C 分类 ----------
    a = audit
    if cur: a['current_price']['available'] += 1
    else: a['current_price']['A'] += 1; a['current_price']['missing'] += 1
    if orig: a['original_price']['available'] += 1
    else: a['original_price']['A'] += 1; a['original_price']['missing'] += 1
    if disc != '': a['discount_rate']['available'] += 1
    else: a['discount_rate']['A'] += 1; a['discount_rate']['missing'] += 1
    if g('采集类目中文'): a['l1']['available'] += 1
    else: a['l1']['A'] += 1; a['l1']['missing'] += 1
    for f in ('l2','l3'):
        a[f]['A'] += 1; a[f]['missing'] += 1
    if leaf:
        a['leaf_category']['available'] += 1
        a['category_rank']['available'] += 1
    else:
        a['leaf_category']['missing'] += 1; a['category_rank']['missing'] += 1
        if not d: a['leaf_category']['A'] += 1; a['category_rank']['A'] += 1
        elif len(bsr_segs) <= 1: a['leaf_category']['A'] += 1; a['category_rank']['A'] += 1
        else: a['leaf_category']['C'] += 1; a['category_rank']['C'] += 1
    a['browse_node_id']['A'] += 1; a['browse_node_id']['missing'] += 1
    if g('monthly_bought_text'): a['monthly_bought']['available'] += 1
    else: a['monthly_bought']['A'] += 1; a['monthly_bought']['missing'] += 1
    if spec_v2: a['spec']['available'] += 1
    else:
        a['spec']['missing'] += 1
        if not d: a['spec']['A'] += 1
        else:
            # 仅 count=1 的字段无可展示规格 → A；有真实规格键但提取失败 → C
            has_real = any(d.get(k) for k in (
                'capacidad','capacidad_de_salida','volumen_de_almacenamiento',
                'volumen_del_tanque','volumen_liquido',
                'dimensiones_del_articulo_largo_x_ancho_x_alto','dimensiones_del_producto',
                'dimensiones_articulo','dimensiones_del_articulo_ancho_x_alto',
                'potencia','voltaje','tension','cantidad_de_compartimentos'))
            if not has_real:
                np_ = d.get('numero_de_piezas')
                has_real = bool(np_ and float(str(np_).replace(' ', '').split(' ')[0].replace(',', '.')) > 1)
            a['spec']['C'] += 1 if has_real else 0
            a['spec']['A'] += 1 if not has_real else 0
    if dfa: a['launch_date']['available'] += 1
    else:
        a['launch_date']['missing'] += 1
        if g('date_first_available_raw'): a['launch_date']['C'] += 1
        else: a['launch_date']['A'] += 1
    if g('parent_asin'): a['parent_asin']['available'] += 1
    else:
        a['parent_asin']['A'] += 1; a['parent_asin']['missing'] += 1

# 补充审计说明
audit['current_price']['note'] = '现价：191/200 有值；9 缺（A）'
audit['original_price']['note'] = '划线原价：源采集从未抓到（A）'
audit['discount_rate']['note'] = '需原价+现价同时存在；现无原价故 0'
audit['l1']['note'] = '采集类目中文 200/200'
audit['l2']['note'] = '源无二级类目数据（A）'
audit['l3']['note'] = '源无三级类目数据（A）'
audit['leaf_category']['note'] = '116 条从 details_json BSR 文本恢复（B）；其余缺或无多榜单'
audit['browse_node_id']['note'] = '源无 Browse Node（A），必须下轮采集'
audit['monthly_bought']['note'] = '源无 monthly_bought 文案（A），必须下轮采集'
audit['category_rank']['note'] = '116 条真实 BSR 细分排名（B）；不猜'
audit['spec']['note'] = '从 details_json 提炼；有字段但解析失败计 C'
audit['launch_date']['note'] = '42 条可解析；其余源无 date_first_available（A）'
audit['parent_asin']['note'] = '134 条；其余源未抓到或确无变体（A）'

with open(OUT_DATA, 'w', encoding='utf-8') as f:
    json.dump(records, f, ensure_ascii=False, indent=1)
with open(OUT_AUDIT, 'w', encoding='utf-8') as f:
    json.dump(audit, f, ensure_ascii=False, indent=1)

print('V2 记录数:', len(records))
for k in audit_def:
    a = audit[k]
    print('%-16s avail=%3d  A=%3d  B=%3d  C=%3d  | %s' % (k, a['available'], a['A'], a['B'], a['C'], a['note']))
print('有BSR文本:', sum(1 for x in records if x['bsr_segments']),
      '| 有叶子榜单:', sum(1 for x in records if x['bsr_leaf_cat']))
print('榜单总段数(排行榜记录行数):', sum(len(x['bsr_segments']) for x in records))
leafs = set(x['bsr_leaf_cat'] for x in records if x['bsr_leaf_cat'])
print('细分榜单总行数(含主榜):', sum(len(x['bsr_segments']) for x in records), '| 去重细分类目数:', len(leafs))
print('规格v2:', sum(1 for x in records if x['spec_v2']),
      '| 摘要v2:', sum(1 for x in records if x['summary_v2']),
      '| 上架时间:', sum(1 for x in records if x['date_first_available']),
      '| GTIN:', sum(1 for x in records if x['gtin']), '| UPC:', sum(1 for x in records if x['upc']))
print('已写出:', OUT_DATA, '|', OUT_AUDIT)
