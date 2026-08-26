# -*- coding: utf-8 -*-
"""
Amazon.es 选品数据库重构 —— 工作簿构建
读取 _selected_data.json（预处理结果）+ _translations.json（若存在，模型翻译），
输出 4 个主 sheet + 1 个翻译对照 sheet：
  1. 选品清单   —— 面向人的主表（图片/序号/名称/价格/类目/排名/规格/摘要/上架/链接/选品状态）
  2. 排行榜记录 —— ASIN × 榜单关系（BSR 主/细分类目、排名、采集时间、榜单 URL 参考）
  3. 后台数据   —— 所有原始爬虫字段（31 列原样 + 派生/清洗列）
  4. 类目规划   —— 由原「产品清单」右侧 38 个类目规划迁移而来
  5. 选品清单中文对照（有翻译时生成）—— 西语→中文 同构对照

不覆盖源文件。不重新抓取 Amazon。
"""
import openpyxl
import json
import os
import datetime
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image
from io import BytesIO

OUTDIR = r"E:\amazon_es\.worktrees\reconnaissance\AmazonESBestseller\outputs\amazon_es_catalog_20260825"
SRC_XLSX = os.path.join(OUTDIR, "AmazonES_产品清单与提取信息.xlsx")
OUT_XLSX = os.path.join(OUTDIR, "AmazonES_产品清单与提取信息_选品优化版.xlsx")
DATA_JSON = os.path.join(OUTDIR, "_selected_data.json")
TRANSL_JSON = os.path.join(OUTDIR, "_translations.json")

CAT_URL = {
    '家居与厨房': 'https://www.amazon.es/gp/bestsellers/kitchen/',   # 已核实
    'DIY及工具': '',                                                # 待核实，不猜测
}

# ---------- 样式 ----------
F_TITLE = Font(name='Microsoft YaHei', size=14, bold=True)
F_NOTE = Font(name='Microsoft YaHei', size=9, color='808080')
F_HEAD = Font(name='Microsoft YaHei', size=10, bold=True, color='FFFFFF')
FILL_HEAD = PatternFill('solid', fgColor='2F5496')
FILL_LABEL = PatternFill('solid', fgColor='DDEBF7')
F_BODY = Font(name='Microsoft YaHei', size=10)
F_LINK = Font(name='Microsoft YaHei', size=9, color='0563C1', underline='single')
THIN = Side(style='thin', color='BFBFBF')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
AL_R = Alignment(horizontal='right', vertical='center', wrap_text=True)
AL_T = Alignment(horizontal='left', vertical='top', wrap_text=True)
AL_CT = Alignment(horizontal='center', vertical='top', wrap_text=True)


def write_header(ws, row, headers, widths, fill=True):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = F_HEAD if fill else Font(name='Microsoft YaHei', size=10, bold=True)
        if fill:
            c.fill = FILL_HEAD
        c.alignment = AL_C
        c.border = BORDER
    ws.row_dimensions[row].height = 28
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def style_data(ws, first_row, last_row, ncols, wrap_cols=(), center_cols=(), right_cols=(), link_cols=()):
    for r in range(first_row, last_row + 1):
        for ci in range(1, ncols + 1):
            c = ws.cell(row=r, column=ci)
            if c.font is None or c.font.name != 'Microsoft YaHei':
                c.font = F_BODY
            c.border = BORDER
            if ci in wrap_cols:
                c.alignment = AL_T
            elif ci in center_cols:
                c.alignment = AL_C
            elif ci in right_cols:
                c.alignment = AL_R
            else:
                c.alignment = AL_L
            if ci in link_cols and c.value:
                c.hyperlink = c.value
                c.font = F_LINK
                c.alignment = AL_L


# ---------- 读取数据 ----------
recs = json.load(open(DATA_JSON, encoding='utf-8'))
transl = None
if os.path.exists(TRANSL_JSON):
    transl = json.load(open(TRANSL_JSON, encoding='utf-8'))

# 读取源工作簿（用于内嵌图片）
src_wb = openpyxl.load_workbook(SRC_XLSX)
src_pis = src_wb['产品清单']
src_images = list(src_pis._images)

# 校验图片与商品顺序对齐：源「产品清单」D 列 ASIN 应与 recs 顺序一致
src_asins = []
for r in range(6, 6 + len(recs)):
    v = src_pis.cell(row=r, column=4).value
    src_asins.append(str(v).strip() if v else '')
aligned = all(sa == recs[i]['asin'] for i, sa in enumerate(src_asins) if sa)
if not aligned:
    print('警告：产品清单 sheet 的 ASIN 顺序与提取信息不完全一致，图片按顺序对齐可能存在偏差。')

wb = openpyxl.Workbook()

# =========================================================
# Sheet 1 选品清单
# =========================================================
ws1 = wb.active
ws1.title = '选品清单'
S1_HEAD = ['图片', '序号', '商品名称', '当前价格', '原价', '折扣率',
           '一级类目', '二级类目', '三级类目', '细分类目', '月购买量',
           '类目排名', '规格', '商品详情摘要', '上架时间', '商品链接',
           '图片链接', 'ASIN', '选品状态', '研究备注']
S1_WIDTH = [13, 6, 46, 12, 12, 9, 12, 12, 12, 12, 10, 10, 34, 64, 12, 34, 38, 13, 13, 20]

ws1.merge_cells('A1:T1')
ws1['A1'] = 'Amazon.es 畅销商品内部选品数据库 · 选品清单'
ws1['A1'].font = F_TITLE
ws1['A1'].alignment = AL_C
ws1.row_dimensions[1].height = 24

# 顶部统计区（第 2 行）
stats = [
    ('商品数', len(recs)), ('一级类目数', 2), ('细分类目数', '0（待采集）'),
    ('有月购买量商品数', sum(1 for r in recs if r['monthly_bought_text'])),
    ('有折扣商品数', 0), ('数据采集时间', '2026-08-25'),
]
ci = 1
for label, value in stats:
    lc = ws1.cell(row=2, column=ci, value=label)
    lc.font = Font(name='Microsoft YaHei', size=9, bold=True)
    lc.fill = FILL_LABEL
    lc.alignment = AL_C
    lc.border = BORDER
    vc = ws1.cell(row=2, column=ci + 1, value=value)
    vc.font = Font(name='Microsoft YaHei', size=10, bold=True)
    vc.alignment = AL_C
    vc.border = BORDER
    ci += 2
ws1.row_dimensions[2].height = 20

write_header(ws1, 3, S1_HEAD, S1_WIDTH)
ws1.freeze_panes = 'A4'

# 序号按一级类目重排：家居与厨房 1~150，DIY及工具 1~50（原顺序内保持）
seq = {}
for i, r in enumerate(recs):
    cat = r['采集类目中文']
    seq[cat] = seq.get(cat, 0) + 1
    r['_seq'] = seq[cat]

first = 4
last = first + len(recs) - 1
for i, r in enumerate(recs):
    rr = first + i
    cur_price = r['current_price']
    try:
        cur_price_f = float(cur_price.replace(',', '.')) if cur_price else None
    except ValueError:
        cur_price_f = None
    best_rank = int(r['best_rank']) if r['best_rank'] else None

    ws1.cell(row=rr, column=2, value=r['_seq']).alignment = AL_C
    ws1.cell(row=rr, column=3, value=r['title_es'])
    d = ws1.cell(row=rr, column=4, value=cur_price_f)
    d.number_format = '#,##0.00" €"'
    e = ws1.cell(row=rr, column=5, value=None)
    e.number_format = '#,##0.00" €"'
    f = ws1.cell(row=rr, column=6, value=None)
    f.number_format = '0%'
    ws1.cell(row=rr, column=7, value=r['采集类目中文'])
    ws1.cell(row=rr, column=8, value='待补充')
    ws1.cell(row=rr, column=9, value='待补充')
    ws1.cell(row=rr, column=10, value='待补充')
    ws1.cell(row=rr, column=11, value=None)
    l = ws1.cell(row=rr, column=12, value=best_rank)
    l.number_format = '0'
    ws1.cell(row=rr, column=13, value=r['spec_es'])
    ws1.cell(row=rr, column=14, value=r['summary_es'])
    ws1.cell(row=rr, column=15, value=r['date_first_available_norm'] or None)
    ws1.cell(row=rr, column=16, value=r['product_url'])
    ws1.cell(row=rr, column=17, value=r['image_url'])
    ws1.cell(row=rr, column=18, value=r['asin'])
    ws1.cell(row=rr, column=19, value='待评估')
    ws1.cell(row=rr, column=20, value=None)

    # 二级/三级/细分类目：待补充 用灰色斜体弱化
    for col in (8, 9, 10):
        c = ws1.cell(row=rr, column=col)
        c.font = Font(name='Microsoft YaHei', size=9, italic=True, color='A6A6A6')
        c.alignment = AL_C

    ws1.row_dimensions[rr].height = 80

style_data(ws1, first, last, 20,
           wrap_cols=(3, 13, 14, 20),
           center_cols=(2, 7, 8, 9, 10, 11, 12, 15, 18, 19),
           right_cols=(4, 5, 6),
           link_cols=(16, 17))

# 图片列（复制源内嵌图片，统一尺寸）
img_width, img_height = 100, 75
for i, im in enumerate(src_images[: len(recs)]):
    if i >= len(recs):
        break
    try:
        data = im.ref.getvalue()
    except Exception:
        continue
    try:
        new_img = Image(BytesIO(data))
        new_img.width = img_width
        new_img.height = img_height
        ws1.add_image(new_img, 'A%d' % (first + i))
    except Exception as ex:
        print('图片嵌入失败 row %d: %s' % (first + i, ex))
ws1.column_dimensions['A'].width = 14

# 选品状态下拉
dv = DataValidation(type='list', formula1='"待评估,重点关注,暂不考虑,已研究"', allow_blank=True,
                    showDropDown=False)
dv.error = '请从下拉列表中选择'
dv.errorTitle = '无效输入'
dv.prompt = '待评估 / 重点关注 / 暂不考虑 / 已研究'
dv.promptTitle = '选品状态'
ws1.add_data_validation(dv)
dv.add('S%d:S%d' % (first, last))

ws1.auto_filter.ref = 'A3:T%d' % last

# =========================================================
# Sheet 2 排行榜记录
# =========================================================
ws2 = wb.create_sheet('排行榜记录')
S2_HEAD = ['asin', '采集类目中文', '采集类目西语', '榜单主类目(来自BSR)', '榜单细分类目(来自BSR)',
           '细分榜单排名(来自BSR)', '类目排名(榜单)', '排名次数', '首次采集时间', '最后采集时间',
           '榜单URL(参考)', '备注']
S2_WIDTH = [14, 12, 18, 22, 26, 12, 12, 10, 20, 20, 40, 26]
write_header(ws2, 1, S2_HEAD, S2_WIDTH)
ws2.freeze_panes = 'A2'

for i, r in enumerate(recs):
    rr = 2 + i
    cat_url = CAT_URL.get(r['采集类目中文'], '')
    note = ''
    if r['ranking_count'] and int(r['ranking_count']) > 1:
        note = '上榜 %s 次，仅记录最佳排名' % r['ranking_count']
    if not r['bsr_main_cat'] and r['details_json']:
        note = (note + '；' if note else '') + 'BSR 类目未解析'
    if not cat_url:
        note = (note + '；' if note else '') + '榜单 URL 待补充'
    ws2.cell(row=rr, column=1, value=r['asin'])
    ws2.cell(row=rr, column=2, value=r['采集类目中文'])
    ws2.cell(row=rr, column=3, value=r['采集类目西语'])
    ws2.cell(row=rr, column=4, value=r['bsr_main_cat'] or None)
    ws2.cell(row=rr, column=5, value=r['bsr_leaf_cat'] or None)
    ws2.cell(row=rr, column=6, value=int(r['bsr_leaf_rank']) if r['bsr_leaf_rank'] else None)
    ws2.cell(row=rr, column=7, value=int(r['best_rank']) if r['best_rank'] else None)
    ws2.cell(row=rr, column=8, value=int(r['ranking_count']) if r['ranking_count'] else None)
    ws2.cell(row=rr, column=9, value=datetime.datetime.fromisoformat(r['first_seen_dt']) if r['first_seen_dt'] else None)
    ws2.cell(row=rr, column=10, value=datetime.datetime.fromisoformat(r['last_seen_dt']) if r['last_seen_dt'] else None)
    ws2.cell(row=rr, column=11, value=cat_url or None)
    ws2.cell(row=rr, column=12, value=note or None)

last2 = 1 + len(recs)
style_data(ws2, 2, last2, 12,
           wrap_cols=(12,),
           center_cols=(1, 2, 6, 7, 8),
           link_cols=(11,))
for r in range(2, last2 + 1):
    ws2.cell(row=r, column=9).number_format = 'yyyy-mm-dd hh:mm'
    ws2.cell(row=r, column=10).number_format = 'yyyy-mm-dd hh:mm'
    ws2.cell(row=r, column=7).number_format = '0'
    ws2.cell(row=r, column=6).number_format = '0'
    ws2.cell(row=r, column=8).number_format = '0'
ws2.auto_filter.ref = 'A1:L%d' % last2

# =========================================================
# Sheet 3 后台数据
# =========================================================
ws3 = wb.create_sheet('后台数据')
S3_HEAD = ['采集类目中文', '采集类目西语', '详情已提取', '图片已下载', 'asin', 'parent_asin',
           'title_es', 'brand', 'price', 'original_price', 'current_price', 'currency',
           'discount_rate', 'rating', 'review_count', 'monthly_bought_text', 'image_url',
           'product_url', 'details_json', 'details', 'specification', 'date_first_available',
           'date_first_available_raw', 'first_seen', 'last_seen', 'ranking_count', 'best_rank',
           'image_path', 'image_download_status', 'image_download_error', 'parent_asin_status',
           'brand_raw', '详情状态', '图片状态', 'monthly_bought_raw', 'monthly_bought_min']
S3_WIDTH = [13, 16, 10, 10, 14, 14, 50, 18, 12, 12, 12, 9, 12, 20, 12, 14, 40, 34,
            70, 50, 16, 16, 18, 18, 18, 10, 10, 22, 18, 18, 16, 18, 10, 10, 16, 16]
write_header(ws3, 1, S3_HEAD, S3_WIDTH)
ws3.freeze_panes = 'A2'

for i, r in enumerate(recs):
    rr = 2 + i
    vals = [r['采集类目中文'], r['采集类目西语'], r['详情已提取'], r['图片已下载'], r['asin'],
            r['parent_asin'], r['title_es'], r['brand'], r['price'], r['original_price'],
            r['current_price'], r['currency'], r['discount_rate'], r['rating'],
            r['review_count'], r['monthly_bought_text'], r['image_url'], r['product_url'],
            r['details_json'], r['details'], r['specification'], r['date_first_available_norm'] or None,
            r['date_first_available_raw'],
            datetime.datetime.fromisoformat(r['first_seen_dt']) if r['first_seen_dt'] else None,
            datetime.datetime.fromisoformat(r['last_seen_dt']) if r['last_seen_dt'] else None,
            r['ranking_count'], r['best_rank'], r['image_path'], r['image_download_status'],
            r['image_download_error'], r['parent_asin_status'], r['brand_raw'],
            r['详情状态'], r['图片状态'], None, None]
    for ci, v in enumerate(vals, 1):
        ws3.cell(row=rr, column=ci, value=v)

last3 = 1 + len(recs)
style_data(ws3, 2, last3, 36,
           wrap_cols=(7, 19, 20),
           center_cols=(1, 2, 3, 4, 5, 6, 12, 13, 14, 15, 26, 27, 29, 31, 33, 34),
           link_cols=(17, 18))
for r in range(2, last3 + 1):
    ws3.cell(row=r, column=24).number_format = 'yyyy-mm-dd hh:mm'
    ws3.cell(row=r, column=25).number_format = 'yyyy-mm-dd hh:mm'
    ws3.cell(row=r, column=22).number_format = 'yyyy-mm-dd'
ws3.auto_filter.ref = 'A1:AJ%d' % last3

# =========================================================
# Sheet 4 类目规划
# =========================================================
ws4 = wb.create_sheet('类目规划')
S4_HEAD = ['#', '中文一级类目', 'Amazon 西语名称', '建议', '我的判断']
S4_WIDTH = [5, 22, 26, 10, 60]
write_header(ws4, 1, S4_HEAD, S4_WIDTH)
ws4.freeze_panes = 'A2'

for rr in range(3, 41):  # 源 P3:T40 = 38 行
    src_row = rr
    p = src_pis.cell(row=src_row, column=16).value
    q = src_pis.cell(row=src_row, column=17).value
    rcol = src_pis.cell(row=src_row, column=18).value
    s = src_pis.cell(row=src_row, column=19).value
    t = src_pis.cell(row=src_row, column=20).value
    dst = rr - 1
    vals = [p, q, rcol, s, t]
    for ci, v in enumerate(vals, 1):
        ws4.cell(row=dst, column=ci, value=v)

last4 = 39
style_data(ws4, 2, last4, 5, wrap_cols=(5,), center_cols=(1, 4))
ws4.auto_filter.ref = 'A1:E%d' % last4

# =========================================================
# Sheet 5 选品清单中文对照（有翻译时）
# =========================================================
if transl:
    ws5 = wb.create_sheet('选品清单中文对照')
    S5_HEAD = ['图片', '序号', '商品名称(西语)', '商品名称(中文)', '当前价格', '原价', '折扣率',
               '一级类目', '二级类目', '三级类目', '细分类目', '月购买量', '类目排名',
               '规格(西语)', '规格(中文)', '商品详情摘要(西语)', '商品详情摘要(中文)',
               '上架时间', '商品链接', '图片链接', 'ASIN', '选品状态', '研究备注']
    S5_WIDTH = [13, 6, 46, 46, 12, 12, 9, 12, 12, 12, 12, 10, 10, 34, 34, 60, 60,
                12, 34, 38, 13, 13, 20]
    ws5.merge_cells('A1:W1')
    ws5['A1'] = '选品清单 · 西语→中文对照翻译（翻译由模型生成，供内部评审参考）'
    ws5['A1'].font = F_TITLE
    ws5['A1'].alignment = AL_C
    ws5.row_dimensions[1].height = 24
    ws5.merge_cells('A2:W2')
    ws5['A2'] = '说明：本表为「选品清单」的同构对照版，文本字段并排显示西语原文与中文译文，行序与选品清单一一对应。'
    ws5['A2'].font = F_NOTE
    ws5['A2'].alignment = AL_L
    ws5.row_dimensions[2].height = 18
    write_header(ws5, 3, S5_HEAD, S5_WIDTH)
    ws5.freeze_panes = 'A4'

    for i, r in enumerate(recs):
        rr = 4 + i
        # 翻译按 row_idx 对应（重复 ASIN 的两行内容可能不同）
        tr = transl.get(str(r['row_idx']), {})
        cur_price_f = None
        if r['current_price']:
            try:
                cur_price_f = float(r['current_price'].replace(',', '.'))
            except ValueError:
                cur_price_f = None
        ws5.cell(row=rr, column=2, value=r['_seq'])
        ws5.cell(row=rr, column=3, value=r['title_es'])
        ws5.cell(row=rr, column=4, value=tr.get('title_zh') or '')
        d = ws5.cell(row=rr, column=5, value=cur_price_f)
        d.number_format = '#,##0.00" €"'
        ws5.cell(row=rr, column=6, value=None).number_format = '#,##0.00" €"'
        ws5.cell(row=rr, column=7, value=None).number_format = '0%'
        ws5.cell(row=rr, column=8, value=r['采集类目中文'])
        for col in (9, 10, 11):
            c = ws5.cell(row=rr, column=col, value='待补充')
            c.font = Font(name='Microsoft YaHei', size=9, italic=True, color='A6A6A6')
        ws5.cell(row=rr, column=12, value=None)
        ws5.cell(row=rr, column=13, value=int(r['best_rank']) if r['best_rank'] else None)
        ws5.cell(row=rr, column=14, value=r['spec_es'])
        ws5.cell(row=rr, column=15, value=tr.get('spec_zh') or '')
        ws5.cell(row=rr, column=16, value=r['summary_es'])
        ws5.cell(row=rr, column=17, value=tr.get('summary_zh') or '')
        ws5.cell(row=rr, column=18, value=r['date_first_available_norm'] or None)
        ws5.cell(row=rr, column=19, value=r['product_url'])
        ws5.cell(row=rr, column=20, value=r['image_url'])
        ws5.cell(row=rr, column=21, value=r['asin'])
        ws5.cell(row=rr, column=22, value='待评估')
        ws5.cell(row=rr, column=23, value=None)
        ws5.row_dimensions[rr].height = 80

    style_data(ws5, 4, 3 + len(recs), 23,
               wrap_cols=(3, 4, 14, 15, 16, 17, 23),
               center_cols=(2, 8, 9, 10, 11, 12, 13, 18, 21, 22),
               right_cols=(5, 6, 7),
               link_cols=(19, 20))
    for i in range(len(recs)):
        try:
            data = src_images[i].ref.getvalue()
            new_img = Image(BytesIO(data))
            new_img.width = img_width
            new_img.height = img_height
            ws5.add_image(new_img, 'A%d' % (4 + i))
        except Exception:
            pass
    dv5 = DataValidation(type='list', formula1='"待评估,重点关注,暂不考虑,已研究"', allow_blank=True)
    ws5.add_data_validation(dv5)
    dv5.add('V4:V%d' % (3 + len(recs)))
    ws5.auto_filter.ref = 'A3:W%d' % (3 + len(recs))
else:
    print('未找到 _translations.json，跳过「选品清单中文对照」sheet（后续生成翻译后重新运行即可补齐）。')

# ---------- 保存 ----------
wb.save(OUT_XLSX)
print('已生成:', OUT_XLSX)
print('sheets:', wb.sheetnames)
