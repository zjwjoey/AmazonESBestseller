# -*- coding: utf-8 -*-
"""Excel 工作簿导出——B3x 新契约（DATA_MODEL §20-§23 / QA_RULES §21-§25/§29）。

默认工作簿 3 张表（§20），顺序固定：
    1. 类目规划
    2. 西班牙语选品清单
    3. 中文选品清单

契约要点：
  - 中文表冻结 26 列（§21 精确顺序）；西语表 25 列（§23 建议对应内容）；
  - ``备注`` 按 ASIN 从前版工作簿合并保留（§22），自动流程不清空/不覆盖；
  - 中西表 ASIN 集一致、确定性排序（§24，ASIN 升序）；
  - 图片按 ASIN 锚定内嵌于中文表 ``01 图片``（§21），西语表不嵌图；
  - exporter 只渲染 canonical 字段，不猜测品牌/类目/排名/类型/规格（§25）；
  - 完整商品详情/商品卖点：``product_details_es/zh``、``feature_bullets_es/zh``
    由 pipeline 渲染（translation/full_detail），无原始详情数据时留空不臆造（§29）；
  - 核心规格（西语）/Parent ASIN 缺失时**留空不臆造**（§29 缺失不自动失败）。
"""
from __future__ import annotations

from typing import Dict, Iterable, List, Mapping, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------- 样式常量（V2 头部 + V1 字体/对齐） ----------
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
F_BODY = Font(size=10)
F_LINK = Font(color='0563C1', underline='single', size=9)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')
AL_C = Alignment(horizontal='center', vertical='center', wrap_text=True)
AL_L = Alignment(horizontal='left', vertical='center', wrap_text=True)
AL_R = Alignment(horizontal='right', vertical='center', wrap_text=True)

CURRENCY_FMT = '#,##0.00" €"'
PERCENT_FMT = '0%'

#: 中文选品清单——冻结 26 列（DATA_MODEL §21 精确顺序，QA_RULES §23）
HEAD_ZH = [
    '图片', '序号', 'ASIN', 'Parent ASIN', '商品名称（中文）', '品牌',
    '当前售价', '划线原价', '折扣率', '评分', '评论数', '月购买量',
    '一级类目', '二级类目', '三级类目', '细分类目', '畅销榜排名',
    '当前选中规格 / 变体', '核心规格（中文）', '完整商品详情（中文）',
    '商品卖点（中文）', '首次上架日期', '卖家', '商品链接', '图片链接', '备注',
]
WIDTH_ZH = [10, 6, 14, 14, 40, 16, 10, 10, 8, 8, 9, 10,
            13, 13, 13, 20, 10, 24, 26, 38, 38, 12, 18, 26, 32, 18]

#: 西班牙语选品清单——25 列（DATA_MODEL §23 建议对应内容）
HEAD_ES = [
    '序号', 'ASIN', 'Parent ASIN', '商品名称（西语）', '品牌',
    '当前售价', '划线原价', '折扣率', '评分', '评论数', '月购买量',
    '一级类目', '二级类目', '三级类目', '细分类目', '畅销榜排名',
    '当前选中规格 / 变体（西语）', '核心规格（西语）',
    '完整商品详情（西语原文）', '商品卖点（西语原文）',
    '首次上架日期', '卖家', '商品链接', '图片链接', '备注',
]
WIDTH_ES = [6, 14, 14, 40, 16, 10, 10, 8, 8, 9, 10,
            13, 13, 13, 20, 10, 24, 26, 38, 38, 12, 18, 26, 32, 18]


def write_header(ws, row: int, headers: Iterable[str], widths: Iterable[int]) -> None:
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = HDR_FONT
        cell.fill = HDR_FILL
        cell.alignment = AL_C
        cell.border = BORDER
    ws.row_dimensions[row].height = 28
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def to_num(v) -> Optional[float]:
    """西语/欧元文本 → float；空或无法解析 → None（§29 缺失不臆造）。"""
    if v is None or str(v).strip() == '':
        return None
    try:
        return float(str(v).replace(',', '.').replace('€', '').strip())
    except ValueError:
        return None


def _notes_of(rec: Mapping) -> str:
    """备注：canonical 字段 ``notes``，中文标签 ``备注``（DATA_MODEL §19）。"""
    for k in ('备注', 'notes'):
        v = rec.get(k)
        if v not in (None, ''):
            return str(v)
    return ''


def _title_zh(rec: Mapping, translations: Optional[Mapping]) -> str:
    """商品名称（中文）：优先记录已有 title_zh，否则回填翻译表（不覆盖）。"""
    v = rec.get('title_zh')
    if v not in (None, ''):
        return str(v)
    asin = rec.get('asin')
    tr = (translations or {}).get(asin) if asin else None
    if isinstance(tr, dict):
        return tr.get('title_zh') or ''
    return ''


# ---------- 行值（只读 canonical 字段，渲染不改源记录） ----------

def _zh_values(rec: Mapping, seq: int, translations: Optional[Mapping]) -> List:
    """中文选品清单一行（26 列，§22 映射）。"""
    return [
        None,                                        # 01 图片（按 ASIN 内嵌，§21）
        seq,                                         # 02 序号
        rec.get('asin'),                             # 03 ASIN
        rec.get('parent_asin'),                      # 04 Parent ASIN
        _title_zh(rec, translations),                # 05 商品名称（中文）
        rec.get('brand') or '',                      # 06 品牌
        to_num(rec.get('current_price')),            # 07 当前售价
        to_num(rec.get('original_price')),           # 08 划线原价
        to_num(rec.get('discount_rate')),            # 09 折扣率
        rec.get('rating') or '',                     # 10 评分
        rec.get('review_count') or '',               # 11 评论数
        rec.get('monthly_bought_min') or '',         # 12 月购买量
        rec.get('category_l1') or '',                # 13 一级类目
        rec.get('category_l2') or '',                # 14 二级类目
        rec.get('category_l3') or '',                # 15 三级类目
        rec.get('leaf_category') or '',              # 16 细分类目
        rec.get('bestseller_rank') or '',            # 17 畅销榜排名
        rec.get('selected_variation_raw') or '',     # 18 当前选中规格 / 变体
        rec.get('spec_v2') or '',                    # 19 核心规格（中文）
        rec.get('product_details_zh') or '',         # 20 完整商品详情（中文）
        rec.get('feature_bullets_zh') or '',         # 21 商品卖点（中文）
        rec.get('date_first_available') or '',       # 22 首次上架日期
        rec.get('seller') or rec.get('seller_raw') or '',  # 23 卖家
        rec.get('product_url') or '',                # 24 商品链接
        rec.get('image_url') or '',                  # 25 图片链接
        _notes_of(rec),                              # 26 备注
    ]


def _es_values(rec: Mapping, seq: int) -> List:
    """西班牙语选品清单一行（25 列，§23 对应内容）。"""
    return [
        seq,                                         # 序号
        rec.get('asin'),                             # ASIN
        rec.get('parent_asin'),                      # Parent ASIN
        rec.get('title_es_raw') or '',               # 商品名称（西语）
        rec.get('brand') or '',                      # 品牌
        to_num(rec.get('current_price')),            # 当前售价
        to_num(rec.get('original_price')),           # 划线原价
        to_num(rec.get('discount_rate')),            # 折扣率
        rec.get('rating') or '',                     # 评分
        rec.get('review_count') or '',               # 评论数
        rec.get('monthly_bought_min') or '',         # 月购买量
        rec.get('category_l1') or '',                # 一级类目
        rec.get('category_l2') or '',                # 二级类目
        rec.get('category_l3') or '',                # 三级类目
        rec.get('leaf_category') or '',              # 细分类目
        rec.get('bestseller_rank') or '',            # 畅销榜排名
        rec.get('selected_variation_raw') or '',     # 当前选中规格 / 变体（西语）
        '',                                          # 核心规格（西语）→ 规格派生，暂留空
        rec.get('product_details_es') or '',         # 完整商品详情（西语原文）
        rec.get('feature_bullets_es') or '',         # 商品卖点（西语原文）
        rec.get('date_first_available') or '',       # 首次上架日期
        rec.get('seller') or rec.get('seller_raw') or '',  # 卖家
        rec.get('product_url') or '',                # 商品链接
        rec.get('image_url') or '',                  # 图片链接
        _notes_of(rec),                              # 备注
    ]


def _write_product_sheet(ws, headers, widths, records, value_fn) -> Dict[str, int]:
    """表头在第 1 行，数据自第 2 行；返回 {ASIN: 1-based 行号}。"""
    write_header(ws, 1, headers, widths)
    row_of_asin: Dict[str, int] = {}
    for i, rec in enumerate(records):
        row = 2 + i
        asin = rec.get('asin')
        if asin:
            row_of_asin[str(asin).upper()] = row
        for ci, v in enumerate(value_fn(rec, i + 1), 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.border = BORDER
            cell.font = F_BODY
            cell.alignment = WRAP
        ws.row_dimensions[row].height = 70
    return row_of_asin


def _style_sheet(ws, nrows: int, price_cols=(), percent_col: int = 0,
                 link_cols=(), center_cols=()) -> None:
    for r in range(2, 2 + nrows):
        for ci in range(1, ws.max_column + 1):
            cell = ws.cell(row=r, column=ci)
            if ci in price_cols and isinstance(cell.value, (int, float)):
                cell.number_format = CURRENCY_FMT
            if ci == percent_col and isinstance(cell.value, (int, float)):
                cell.number_format = PERCENT_FMT
            if ci in center_cols:
                cell.alignment = AL_C
            if ci in link_cols and cell.value:
                cell.hyperlink = cell.value
                cell.font = F_LINK


def embed_images_by_asin(ws, images_by_asin: Mapping, row_of_asin: Mapping,
                         col: str = 'A') -> None:
    """按 ASIN 锚定内嵌图片（QA_RULES §21/§51-§52），不按位置。

    images_by_asin: {ASIN: (BytesIO|bytes, width, height)}；row_of_asin: {ASIN: 1-based 行号}。
    """
    for asin, (data, width, height) in images_by_asin.items():
        row = row_of_asin.get(str(asin).upper())
        if row is None:
            continue
        try:
            img = XLImage(data)
        except Exception:
            continue
        img.width = width or 70
        img.height = height or 70
        ws.add_image(img, '%s%d' % (col, row))


def _find_header_row(ws, required_col: str) -> Optional[int]:
    for r in range(1, min(ws.max_row, 10) + 1):
        row_vals = [ws.cell(row=r, column=ci).value for ci in range(1, ws.max_column + 1)]
        if any(v is not None and str(v).strip() == required_col for v in row_vals):
            return r
    return None


def _col_index(header: List[str], name: str) -> Optional[int]:
    for i, h in enumerate(header):
        if h == name:
            return i
    return None


def merge_manual_fields(records: List[Mapping], prev_workbook=None) -> List[Mapping]:
    """从前版工作簿按 ASIN 合并 ``备注``（QA_RULES §22，DATA_MODEL §19）。

    只回填前版非空备注；自动流程不清空、不覆盖、不以 QA 状态替换。
    找任意同时含 ``ASIN`` + ``备注`` 列的 sheet（新契约：中西选品清单）。
    """
    out = [dict(r) for r in records]
    if prev_workbook is None:
        return out
    for name in getattr(prev_workbook, 'sheetnames', ()):
        ws = prev_workbook[name]
        hr = _find_header_row(ws, 'ASIN')
        if hr is None:
            continue
        header = [str(c.value).strip() if c.value is not None else '' for c in ws[hr]]
        ci_asin = _col_index(header, 'ASIN')
        ci_notes = _col_index(header, '备注')
        if ci_asin is None or ci_notes is None:
            continue
        manual = {}
        for r in ws.iter_rows(min_row=hr + 1, values_only=True):
            if ci_asin >= len(r) or r[ci_asin] is None or str(r[ci_asin]).strip() == '':
                continue
            a = str(r[ci_asin]).strip().upper()
            if ci_notes < len(r) and r[ci_notes] not in (None, ''):
                manual[a] = str(r[ci_notes]).strip()
        if not manual:
            continue
        for rec in out:
            a = str(rec.get('asin', '')).strip().upper()
            if a in manual and _notes_of(rec) == '':
                rec['备注'] = manual[a]
        break
    return out


def export_workbook(records: List[Mapping],
                    translations: Optional[Mapping] = None,
                    images_by_asin: Optional[Mapping] = None,
                    category_planning: Optional[List] = None,
                    prev_workbook=None,
                    out_path=None,
                    collected_at: Optional[str] = None) -> openpyxl.Workbook:
    """构建默认 3-sheet 契约工作簿（DATA_MODEL §20-§23）并返回；仅当给 out_path 时保存。

    records: 合并后的商品表记录（内部拷贝 + ASIN 升序确定性排序，§24）；
    translations 按 ASIN → {title_zh,...}（回填 商品名称（中文），记录已含则不覆盖）；
    images_by_asin 按 ASIN → (data, w, h)，仅内嵌到中文表 ``01 图片``（§21）；
    category_planning 为 dict 行或 2D 行列表（类目规划，人工维护）。
    本函数不修改传入 records。
    """
    records = [dict(r) for r in records]
    if prev_workbook is not None:
        records = merge_manual_fields(records, prev_workbook)
    # 确定性排序（QA_RULES §24）：ASIN 升序，中西两表同一顺序
    records.sort(key=lambda r: str(r.get('asin') or '').upper())

    wb = openpyxl.Workbook()

    # ---------- Sheet1 类目规划（人工维护） ----------
    ws_cat = wb.active
    ws_cat.title = '类目规划'
    cat_rows = category_planning or []
    cat_header = list(cat_rows[0].keys()) if cat_rows and isinstance(cat_rows[0], dict) \
        else ['#', '中文一级类目', 'Amazon 西语名称', '建议', '我的判断']
    write_header(ws_cat, 1, cat_header, [8, 22, 26, 22, 22])
    for i, r in enumerate(cat_rows):
        if isinstance(r, dict):
            for ci, h in enumerate(cat_header, 1):
                v = r.get(h)
                if v is None:
                    continue
                cell = ws_cat.cell(row=2 + i, column=ci, value=v)
                cell.border = BORDER
                cell.font = F_BODY
                cell.alignment = WRAP
        else:
            for ci, v in enumerate(r, 1):
                if v is None:
                    continue
                cell = ws_cat.cell(row=2 + i, column=ci, value=v)
                cell.border = BORDER
                cell.font = F_BODY
                cell.alignment = WRAP
    ws_cat.freeze_panes = 'A2'

    # ---------- Sheet2 西班牙语选品清单（不嵌图，§21/§23） ----------
    ws_es = wb.create_sheet('西班牙语选品清单')
    _write_product_sheet(ws_es, HEAD_ES, WIDTH_ES, records,
                         lambda rec, seq: _es_values(rec, seq))
    _style_sheet(ws_es, len(records), price_cols=(6, 7), percent_col=8,
                 link_cols=(23, 24), center_cols=(1, 2))
    ws_es.freeze_panes = 'B2'  # 冻结 序号 + 表头
    if records:
        ws_es.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(HEAD_ES)), len(records) + 1)

    # ---------- Sheet3 中文选品清单（26 列冻结 + 按 ASIN 内嵌图片，§21/§23） ----------
    ws_zh = wb.create_sheet('中文选品清单')
    row_of_asin = _write_product_sheet(
        ws_zh, HEAD_ZH, WIDTH_ZH, records,
        lambda rec, seq: _zh_values(rec, seq, translations))
    _style_sheet(ws_zh, len(records), price_cols=(7, 8), percent_col=9,
                 link_cols=(24, 25), center_cols=(2, 3))
    if images_by_asin:
        embed_images_by_asin(ws_zh, images_by_asin, row_of_asin, col='A')
    ws_zh.freeze_panes = 'C2'  # 冻结 01 图片 + 02 序号 + 表头
    if records:
        ws_zh.auto_filter.ref = 'A1:%s%d' % (
            get_column_letter(len(HEAD_ZH)), len(records) + 1)

    if out_path:
        wb.save(out_path)
    return wb
