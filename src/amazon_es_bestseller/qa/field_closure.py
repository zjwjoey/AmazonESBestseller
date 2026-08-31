# -*- coding: utf-8 -*-
"""End-to-end Source → Raw → Canonical → Derived → Display audit.

This module is deliberately read-only.  It diagnoses why an automatic display
field is empty; it never fills a value and never uses a different field as a
guess.  Saved HTML is optional, but when supplied it allows a missing raw value
to be distinguished from a page that did not expose the field.
"""
from __future__ import annotations

from dataclasses import asdict, dataclass
from functools import lru_cache
import json
import re
from pathlib import Path
from typing import Any, Iterable, Mapping, Optional

import openpyxl
from bs4 import BeautifulSoup

from ..export.excel import HEAD_ES, HEAD_ZH, _es_values, _zh_values
from ..models import normalize_asin
from ..normalization.price import discount_rate, parse_price


PASS = "PASS"
SOURCE_MISSING = "SOURCE_MISSING"
PARSER_MISSED = "PARSER_MISSED"
MAPPING_MISSED = "MAPPING_MISSED"
DERIVED_MISSING = "DERIVED_MISSING"
TRANSLATION_INCOMPLETE = "TRANSLATION_INCOMPLETE"
NOT_OBSERVED = "NOT_OBSERVED"
EVIDENCE_UNAVAILABLE = "EVIDENCE_UNAVAILABLE"
EXPORT_MISSING = "EXPORT_MISSING"
EXPORT_VALUE_MISMATCH = "EXPORT_VALUE_MISMATCH"
IMAGE_MISSING = "IMAGE_MISSING"

# These fields are validly empty when Amazon does not expose the corresponding
# page element.  They are coverage signals, not product-data defects.
_CONDITIONAL_FIELDS = frozenset({
    "parent_asin", "brand", "original_price", "discount_rate", "monthly_bought_min",
    "selected_variation_raw", "spec_v2", "product_details_zh", "feature_bullets_zh",
    "date_first_available", "seller",
})


@dataclass(frozen=True)
class FieldClosureResult:
    asin: str
    field: str
    display_column: str
    source_status: str
    raw_status: str
    canonical_status: str
    derived_status: str
    display_status: str
    classification: str
    severity: str
    source_evidence: Any
    raw_evidence: Any
    canonical_value: Any
    derived_value: Any
    display_value: Any
    message: str

    def to_dict(self) -> dict:
        return asdict(self)


# Field names intentionally correspond to the frozen Chinese 26-column contract.
FIELD_COLUMNS = {
    "asin": "ASIN",
    "parent_asin": "Parent ASIN",
    "title_zh": "商品名称（中文）",
    "brand": "品牌",
    "current_price": "当前售价",
    "original_price": "划线原价",
    "discount_rate": "折扣率",
    "rating": "评分",
    "review_count": "评论数",
    "monthly_bought_min": "月购买量",
    "category_l1": "一级类目",
    "category_l2": "二级类目",
    "category_l3": "三级类目",
    "leaf_category": "细分类目",
    "bestseller_rank": "畅销榜排名",
    "selected_variation_raw": "当前选中规格 / 变体",
    "spec_v2": "核心规格（中文）",
    "product_details_zh": "完整商品详情（中文）",
    "feature_bullets_zh": "商品卖点（中文）",
    "date_first_available": "首次上架日期",
    "seller": "卖家",
    "product_url": "商品链接",
    "image_url": "图片链接",
}

_FIELD_ORDER = tuple(FIELD_COLUMNS)
_TRANSLATION_FIELDS = {
    "title_zh": ("title_es_raw", "商品名称（中文）"),
    "category_l1_zh": ("category_l1", "一级类目"),
    "category_l2_zh": ("category_l2", "二级类目"),
    "category_l3_zh": ("category_l3", "三级类目"),
    "leaf_category_zh": ("leaf_category", "细分类目"),
    "selected_variation_zh": ("selected_variation_raw", "当前选中规格 / 变体"),
    "specification_zh": ("specification_es", "核心规格（中文）"),
    "product_details_zh": ("product_details_es", "完整商品详情（中文）"),
    "feature_bullets_zh": ("feature_bullets_es", "商品卖点（中文）"),
}
_SPANISH_MARKERS = re.compile(r"\b(?:el|la|los|las|para|con|sin|del|en|un|una|comprados|piezas|tamaño|capacidad|material|color|negro|blanco|acero|plástico)\b", re.I)


def _translation_residual(source: Any, target: Any) -> bool:
    s, t = str(source or "").strip(), str(target or "").strip()
    if not s:
        return False
    if not t:
        return True
    if t.casefold() == s.casefold():
        # Pure model codes, sizes and unit-bearing quantities are validly
        # identical across languages (e.g. L, 500ml, 122cm x 51cm).
        return bool(_SPANISH_MARKERS.search(s))
    # Long Spanish sentences surviving in a Chinese display value are a
    # stronger signal than legitimate model/brand tokens.
    return len(t.split()) >= 5 and bool(_SPANISH_MARKERS.search(t))
_HTML_PATTERNS = {
    "parent_asin": (r"parent[-_ ]?asin", r"asin de padre", r"asin padre"),
    "title_zh": (r"id=[\"']producttitle",),
    "brand": (r"bylineinfo", r"\bmarca\b", r"\bbrand\b"),
    "current_price": (r"coreprice", r"pricetopay", r"a-price"),
    "original_price": (r"(?:corePrice_feature_div|corePriceDisplay_desktop_feature_div)[\s\S]{0,6000}data-a-strike\s*=\s*[\"']true[\"']",),
    "discount_rate": (r"(?:corePrice_feature_div|corePriceDisplay_desktop_feature_div)[\s\S]{0,6000}data-a-strike\s*=\s*[\"']true[\"']",),
    "rating": (r"acrpopover", r"averagecustomerreviews", r"estrellas"),
    "review_count": (r"acrcustomerreviewtext", r"opiniones de clientes"),
    "monthly_bought_min": (r"comprados el mes pasado", r"comprado[s]? el último mes"),
    "category_l1": (r"breadcrumb", r"zgbs", r"browse node", r"browse_node"),
    "category_l2": (r"breadcrumb", r"zgbs", r"browse node", r"browse_node"),
    "category_l3": (r"breadcrumb", r"zgbs", r"browse node", r"browse_node"),
    "leaf_category": (r"breadcrumb", r"zgbs", r"browse node", r"browse_node"),
    "bestseller_rank": (r"best sellers rank", r"más vendidos", r"n\.?º\s*[\d.,]+\s+en"),
    "selected_variation_raw": (r"id\s*=\s*[\"']variation_name[\"']",
                                r"id\s*=\s*[\"']twister-plus-name-feature[\"']"),
    "spec_v2": (r"productdetails", r"proddetails", r"capacidad", r"dimensiones", r"tamaño"),
    "product_details_zh": (r"productdetails", r"proddetails", r"technical details", r"características"),
    "feature_bullets_zh": (r"feature-bullets", r"about this item", r"acerca de este producto"),
    "date_first_available": (r"fecha de primera disponibilidad", r"date first available",
                             r"producto en amazon\.es desde"),
    "seller": (r"merchantinfo", r"sellerprofile", r"vendido por", r"sold by"),
    "product_url": (r"/dp/[a-z0-9]{10}",),
    "image_url": (r"landingimage", r"mainimage", r"imageblock", r"\.jpg", r"\.png"),
}


def _has(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, (str, bytes)):
        return bool(value.strip())
    if isinstance(value, (list, tuple, set, dict)):
        return bool(value)
    return True


def _text(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _attributes(record: Mapping, detail: Mapping) -> list:
    attrs = record.get("attributes")
    if attrs:
        return attrs if isinstance(attrs, list) else []
    attrs = detail.get("attributes")
    return attrs if isinstance(attrs, list) else []


def _attribute_brand(attrs: Iterable[Mapping]) -> str:
    for item in attrs:
        label = _text(item.get("label_raw")).casefold()
        if label in {"marca", "brand"} and _has(item.get("value_raw")):
            return _text(item["value_raw"])
    return ""


def _html_roots(html_dir: Optional[str | Path | Iterable[str | Path]]) -> list[Path]:
    if not html_dir:
        return []
    values = (html_dir,) if isinstance(html_dir, (str, Path)) else tuple(html_dir)
    return [Path(value) for value in values if Path(value).is_dir()]


def _find_html(html_dir: Optional[str | Path | Iterable[str | Path]], asin: str) -> str:
    if not html_dir:
        return ""
    for root in _html_roots(html_dir):
        candidates = [root / f"{asin}.html", root / f"{asin.upper()}.html", root / f"{asin.lower()}.html"]
        for path in candidates:
            if path.is_file():
                try:
                    return path.read_text(encoding="utf-8", errors="ignore")
                except OSError:
                    pass
        for path in sorted(root.rglob("*.html")):
            if asin.casefold() in path.stem.casefold():
                try:
                    return path.read_text(encoding="utf-8", errors="ignore")
                except OSError:
                    return ""
    return ""


def _html_asin(html: str) -> str:
    """Return the page ASIN from saved HTML without relying on its filename."""
    patterns = (
        r'<input[^>]+(?:id|name)=["\']ASIN["\'][^>]+value=["\']([A-Z0-9]{10})["\']',
        r'<input[^>]+value=["\']([A-Z0-9]{10})["\'][^>]+(?:id|name)=["\']ASIN["\']',
        r'/dp/([A-Z0-9]{10})(?:[/?"\']|$)',
    )
    for pattern in patterns:
        match = re.search(pattern, html, re.I)
        if match:
            return normalize_asin(match.group(1))
    return ""


class _LazyHtmlByAsin(dict):
    """ASIN→HTML mapping that indexes paths without loading every page."""
    def __init__(self, paths: dict[str, Path]):
        super().__init__()
        self._paths = paths

    def get(self, key, default=None):
        path = self._paths.get(str(key).upper())
        if path is None:
            return default
        try:
            return path.read_text(encoding="utf-8", errors="ignore")
        except OSError:
            return default


def _html_by_asin(html_dir: Optional[str | Path | Iterable[str | Path]]) -> Mapping[str, str]:
    """Index saved detail HTML by ASIN while keeping page bodies lazy."""
    indexed: dict[str, Path] = {}
    for root in _html_roots(html_dir):
        for path in sorted(root.rglob("*.html")):
            try:
                html = path.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
            asin = _html_asin(html)
            if not asin and re.fullmatch(r"[A-Z0-9]{10}", path.stem, re.I):
                asin = path.stem.upper()
            if asin and asin not in indexed:
                indexed[asin] = path
    return _LazyHtmlByAsin(indexed)


def _read_ranking_html(ranking_html_dir: Optional[str | Path | Iterable[str | Path]]) -> str:
    """Read ranking-page HTML only; detail-page breadcrumbs are not ranking evidence."""
    chunks = []
    for root in _html_roots(ranking_html_dir):
        for path in sorted(root.rglob("ranking_*.html")):
            try:
                chunks.append(path.read_text(encoding="utf-8", errors="ignore"))
            except OSError:
                continue
    return "\n".join(chunks)


def _html_parent_values(html: str) -> list[str]:
    """Return explicit parent-ASIN values exposed in detail-page markup.

    Amazon commonly includes a hidden ``parentASIN`` input even for a
    standalone product, with the value equal to the child ASIN.  That
    self-reference is not evidence of a variation family and must not make a
    missing ``parent_asin`` look like a parser defect.
    """
    patterns = (
        r'<input[^>]+(?:id|name)=["\']parentASIN["\'][^>]+value=["\']([A-Z0-9]{10})["\']',
        r'<input[^>]+value=["\']([A-Z0-9]{10})["\'][^>]+(?:id|name)=["\']parentASIN["\']',
        r'["\']parentAsin["\']\s*:\s*["\']([A-Z0-9]{10})["\']',
    )
    values: list[str] = []
    for pattern in patterns:
        values.extend(m.upper() for m in re.findall(pattern, html, re.I))
    return list(dict.fromkeys(values))


def _html_visible_evidence(field: str, html: str) -> bool:
    """Check field-specific visible markup instead of generic ID/CSS noise."""
    soup = _cached_soup(html)
    if field == "seller":
        selectors = ("#merchantInfoFeature_feature_div", "#sellerProfileTriggerId",
                     '#tabular-buybox .tabular-buybox-text[role="text"]')
        return any(re.search(r"(?:vendido\s+por|vendedor|remitente\s*/\s*vendedor|sold\s+by)\s+\S+",
                             el.get_text(" ", strip=True), re.I)
                   for sel in selectors for el in soup.select(sel))
    if field == "brand":
        byline = soup.select_one("#bylineInfo")
        if byline is not None and byline.get_text(" ", strip=True):
            return True
        return any(_clean_text_pair(row) for row in soup.select("#productOverview_feature_div tr"))
    if field == "current_price":
        return any(el.get_text(" ", strip=True) for el in soup.select(
            "#corePrice_feature_div .a-price .a-offscreen, "
            "#corePriceDisplay_desktop_feature_div .a-price .a-offscreen, "
            ".apex-pricetopay-value .a-offscreen, .priceToPay .a-offscreen"))
    if field == "rating":
        root = soup.select_one("#averageCustomerReviews_feature_div")
        return bool(root and re.search(r"\d[\d.,]*\s+de\s+5\s+estrellas", root.get_text(" ", strip=True), re.I))
    if field == "review_count":
        el = soup.select_one("#acrCustomerReviewText")
        return bool(el and el.get_text(" ", strip=True))
    if field == "feature_bullets_zh":
        return any(el.select("li") for sel in ("#feature-bullets", "#featurebullets_feature_div", "#pqv-feature-bullets")
                   for el in soup.select(sel))
    return False


@lru_cache(maxsize=8)
def _cached_soup(html: str):
    """Avoid reparsing the same multi-megabyte page once per audited field."""
    return BeautifulSoup(html, "lxml")


def _clean_text_pair(row) -> bool:
    cells = row.select("td")
    if len(cells) < 2:
        return False
    label = re.sub(r"\s+", " ", cells[0].get_text(" ", strip=True)).casefold()
    value = re.sub(r"\s+", " ", cells[1].get_text(" ", strip=True))
    return label in {"marca", "brand"} and bool(value)


def _source_evidence(field: str, record: Mapping, detail: Mapping, ranking: Mapping,
                     html: str) -> tuple[bool, Any]:
    evidence: list[str] = []
    if field == "brand":
        if _has(detail.get("brand_raw")) or _has(record.get("brand_raw")):
            evidence.append("brand_raw")
        if _attribute_brand(_attributes(record, detail)):
            evidence.append("attributes:Marca/Brand")
    elif field.startswith("category_") or field == "leaf_category":
        # A single root category is not evidence that deeper levels exist.  A
        # breadcrumb/path or browse node is evidence for parsing missing levels;
        # a ranking URL alone is only context, not category content.
        if _has(ranking.get(field)) or _has(record.get(field)):
            evidence.append(field)
        if _has(ranking.get("category_path_raw")) or (
                field != "category_l3" and _has(ranking.get("browse_node_id"))):
            evidence.append("ranking.category_path")
        if html and any(re.search(p, html, re.I) for p in _HTML_PATTERNS.get(field, ())):
            evidence.append("html:category_path")
    elif field == "bestseller_rank":
        if _has(ranking.get("bestseller_rank")) or (_has(record.get("bestseller_rank")) and
                                                    (_has(record.get("ranking_source_url")) or _has(record.get("collected_at")))):
            evidence.append("ranking.bestseller_rank")
    elif field in {"title_zh", "brand", "current_price", "original_price", "rating",
                   "review_count", "monthly_bought_min", "parent_asin", "selected_variation_raw",
                   "date_first_available", "seller", "spec_v2", "product_details_zh",
                   "feature_bullets_zh"}:
        # Individual raw evidence is handled below; this branch intentionally does not
        # treat a derived field itself as source evidence.
        pass
    else:
        if _has(detail.get(field)) or _has(record.get(field)):
            evidence.append(field)
    if html:
        if field == "parent_asin":
            child_asin = normalize_asin(record.get("asin"))
            confirmed = [value for value in _html_parent_values(html)
                         if normalize_asin(value) and normalize_asin(value) != child_asin]
            if confirmed:
                evidence.append("html:confirmed_parent_asin")
        elif field in {"seller", "brand", "current_price", "rating", "review_count",
                       "feature_bullets_zh"}:
            if _html_visible_evidence(field, html):
                evidence.append(f"html:visible:{field}")
        elif field in {"original_price", "discount_rate"}:
            # 复用采集器本身的划线价语义，而不是另写一条更宽松的正则：
            # Amazon 的 "Precio único" 用 data-a-strike=true 重述当前售价，
            # 值与现价相同，不是有效划线原价。审计若只匹配 data-a-strike，
            # 会把采集器的正确行为判成 PARSER_MISSED 并拦下合格导出
            # （实采 100 SKU 中 15 例，真实证据 B000LXUWN6）。
            from ..collection.detail import struck_price_from_html
            if struck_price_from_html(html):
                evidence.append(f"html:struck_price:{field}")
        else:
            patterns = _HTML_PATTERNS.get(field, ())
            matched = [p for p in patterns if re.search(p, html, re.I)]
            evidence.extend(f"html:{p}" for p in matched)
    return bool(evidence), evidence


def _raw_value(field: str, record: Mapping, detail: Mapping, ranking: Mapping) -> Any:
    if field == "asin":
        return record.get("asin") or detail.get("asin") or ranking.get("asin")
    if field == "brand":
        return detail.get("brand_raw") or record.get("brand_raw") or _attribute_brand(_attributes(record, detail))
    if field == "title_zh":
        return detail.get("title_es_raw") or record.get("title_es_raw")
    if field in {"current_price", "original_price", "rating", "review_count", "monthly_bought_min"}:
        raw_key = {"current_price": "current_price_raw", "original_price": "original_price_raw",
                   "rating": "rating_raw", "review_count": "review_count_raw",
                   "monthly_bought_min": "monthly_bought_raw"}[field]
        return detail.get(raw_key) or record.get(raw_key)
    if field.startswith("category_") or field == "leaf_category":
        return ranking.get(field) if _has(ranking.get(field)) else record.get(field)
    if field == "bestseller_rank":
        if _has(ranking.get("bestseller_rank")):
            return ranking.get("bestseller_rank")
        if _has(record.get("ranking_source_url")) or _has(record.get("collected_at")):
            return record.get("bestseller_rank")
        return None
    if field in {"spec_v2", "product_details_zh"}:
        if field == "spec_v2":
            # Attribute presence alone is not specification evidence; only a
            # non-empty derived summary can close this field.
            return record.get("spec_v2") or ""
        return _attributes(record, detail) or detail.get("details_json") or record.get("details_json")
    if field == "feature_bullets_zh":
        return detail.get("feature_bullets_raw") or record.get("feature_bullets_raw")
    if field == "date_first_available":
        return detail.get("date_first_available_raw") or record.get("date_first_available_raw")
    if field == "seller":
        return detail.get("seller_raw") or record.get("seller_raw")
    if field == "parent_asin":
        value = detail.get("parent_asin") or record.get("parent_asin_raw")
        # A child ASIN repeated as its own parent is not variation-family evidence.
        if normalize_asin(value) == normalize_asin(record.get("asin")):
            return None
        return value
    if field == "selected_variation_raw":
        return detail.get("selected_variation_raw") or record.get("selected_variation_raw")
    return detail.get(field) or record.get(field)


def _canonical_value(field: str, record: Mapping) -> Any:
    if field == "seller":
        return record.get("seller") or record.get("seller_raw")
    return record.get(field)


def _derived_value(field: str, record: Mapping) -> Any:
    if field == "title_zh":
        return record.get("title_zh")
    if field == "spec_v2":
        return record.get("spec_v2") or record.get("specification")
    if field == "product_details_zh":
        return record.get("product_details_zh")
    if field == "feature_bullets_zh":
        return record.get("feature_bullets_zh")
    if field == "discount_rate":
        return record.get("discount_rate")
    return _canonical_value(field, record)


def _display_value(field: str, record: Mapping) -> Any:
    return _canonical_value(field, record)


def _classify(field: str, source: bool, raw: Any, canonical: Any, derived: Any,
              display: Any, raw_evidence: Any, page_available: bool) -> tuple[str, str, str]:
    if _has(display):
        return PASS, "INFO", "字段已闭环进入展示层"
    if not source:
        if field in _CONDITIONAL_FIELDS:
            if page_available:
                return NOT_OBSERVED, "INFO", "已保存页面未展示该条件字段"
            return EVIDENCE_UNAVAILABLE, "INFO", "没有可核验的详情页证据，无法判断字段是否展示"
        return SOURCE_MISSING, "P2", "当前页面/榜单未发现可靠来源证据"
    if not _has(raw):
        return PARSER_MISSED, "P1", "来源证据存在，但 collector 未保存 raw 字段"
    if not _has(canonical):
        return MAPPING_MISSED, "P1", "raw 证据存在，但未映射到 canonical 字段"
    if not _has(derived) or not _has(display):
        return DERIVED_MISSING, "P1", "raw/canonical 存在，但派生或展示字段为空"
    return PASS, "INFO", "字段已闭环进入展示层"


def _audit_one(field: str, asin: str, record: Mapping, detail: Mapping,
               ranking: Mapping, html: str) -> FieldClosureResult:
    raw = _raw_value(field, record, detail, ranking)

    # Once a raw value is already present, source evidence is established by the
    # raw input below.  Avoid rescanning multi-megabyte saved HTML for every
    # field; HTML inspection is only needed to explain an otherwise empty raw
    # value (the PARSER_MISSED/MAPPING_MISSED path).
    source, source_evidence = _source_evidence(
        field, record, detail, ranking, "" if _has(raw) else html
    )
    if _has(raw) and not source:
        source = True
        source_evidence = ["raw_input"]
    canonical = _canonical_value(field, record)
    derived = _derived_value(field, record)
    display = _display_value(field, record)

    # For translated/derived columns the canonical input is the Spanish/raw
    # canonical value, while the target column is represented by ``derived``.
    # This distinction is what separates DERIVED_MISSING from MAPPING_MISSED.
    if field == "title_zh":
        canonical = detail.get("title_es_raw") or record.get("title_es_raw")
    elif field in {"product_details_zh", "feature_bullets_zh", "spec_v2"}:
        canonical = raw

    # Derived fields have source evidence from their input chain, not from their own
    # display value.  Prices require two valid source values before discount is expected.
    if field == "discount_rate":
        cur = parse_price(detail.get("current_price_raw") or record.get("current_price_raw") or record.get("current_price"))
        orig = parse_price(detail.get("original_price_raw") or record.get("original_price_raw") or record.get("original_price"))
        source = cur is not None and orig is not None and orig > cur
        source_evidence = ["current_price", "valid_struck_original_price"] if source else source_evidence
        raw = {"current_price": cur, "original_price": orig} if cur is not None or orig is not None else ""
    if field in {"title_zh", "spec_v2", "product_details_zh", "feature_bullets_zh"}:
        source = _has(raw) or bool(source_evidence)

    classification, severity, message = _classify(
        field, source, raw, canonical, derived, display, raw_evidence=raw,
        page_available=bool(html))
    if field == "title_zh" and display and str(canonical).casefold() == str(display).casefold():
        classification, severity = TRANSLATION_INCOMPLETE, "P1"
        message = "中文展示字段为空或仍保留西语原文/整句"
    if field == "original_price":
        cur = parse_price(detail.get("current_price_raw") or record.get("current_price_raw") or record.get("current_price"))
        orig = parse_price(raw)
        if cur is not None and orig is not None and orig <= cur:
            classification, severity = "ORIGINAL_PRICE_INVALID", "P1"
            message = "原价小于或等于当前售价，不作为有效划线原价展示"
            canonical = derived = display = None
    return FieldClosureResult(
        asin=asin,
        field=field,
        display_column=FIELD_COLUMNS[field],
        source_status=("present" if source else
                       ("not_observed" if classification == NOT_OBSERVED else
                        "unavailable" if classification == EVIDENCE_UNAVAILABLE else "missing")),
        raw_status="present" if _has(raw) else "missing",
        canonical_status="present" if _has(canonical) else "missing",
        derived_status="present" if _has(derived) else "missing",
        display_status="present" if _has(display) else "missing",
        classification=classification,
        severity=severity,
        source_evidence=source_evidence,
        raw_evidence=raw,
        canonical_value=canonical,
        derived_value=derived,
        display_value=display,
        message=message,
    )


def _export_issue(asin: str, field: str, display_column: str, classification: str,
                  message: str, expected: Any = None, actual: Any = None) -> dict:
    """Represent a workbook-only discrepancy in the same report shape as fields."""
    return FieldClosureResult(
        asin=asin,
        field=field,
        display_column=display_column,
        source_status="present",
        raw_status="present",
        canonical_status="present",
        derived_status="present",
        display_status="missing" if classification == EXPORT_MISSING else "mismatch",
        classification=classification,
        severity="P1",
        source_evidence="export_contract",
        raw_evidence=expected,
        canonical_value=expected,
        derived_value=expected,
        display_value=actual,
        message=message,
    ).to_dict()


def _same_display_value(expected: Any, actual: Any) -> bool:
    """Treat exporter blanks consistently while preserving numeric comparisons."""
    if expected in (None, "") and actual in (None, ""):
        return True
    return expected == actual


def _workbook_images_by_row(ws) -> set[int]:
    rows: set[int] = set()
    for image in getattr(ws, "_images", ()):
        anchor = getattr(image, "anchor", None)
        marker = getattr(anchor, "_from", None)
        row = getattr(marker, "row", None)
        if isinstance(row, int):
            rows.add(row + 1)  # drawing anchors are zero-based; worksheet rows are one-based
    return rows


def _audit_workbook(records: list[Mapping], workbook_path: str | Path,
                    translations: Optional[Mapping] = None) -> list[dict]:
    """Read an exported workbook and reconcile its display layer by ASIN."""
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=False)
    except (OSError, ValueError, openpyxl.utils.exceptions.InvalidFileException) as exc:
        return [_export_issue("", "workbook", "工作簿", EXPORT_MISSING,
                              "无法读取导出工作簿：%s" % type(exc).__name__)]

    findings: list[dict] = []
    ordered = sorted(records, key=lambda rec: normalize_asin(rec.get("asin")))
    sheets = (
        ("西班牙语选品清单", HEAD_ES, lambda rec, seq: _es_values(rec, seq), 2, False),
        ("中文选品清单", HEAD_ZH, lambda rec, seq: _zh_values(rec, seq, translations), 3, True),
    )
    for sheet_name, headers, values_fn, asin_col, image_sheet in sheets:
        if sheet_name not in wb.sheetnames:
            findings.append(_export_issue("", "workbook_sheet", sheet_name, EXPORT_MISSING,
                                          "导出工作簿缺少工作表 %s" % sheet_name))
            continue
        ws = wb[sheet_name]
        actual_headers = [ws.cell(1, col).value for col in range(1, len(headers) + 1)]
        if actual_headers != headers:
            findings.append(_export_issue("", "workbook_header", sheet_name, EXPORT_VALUE_MISMATCH,
                                          "工作表表头与冻结导出契约不一致", headers, actual_headers))
        row_by_asin: dict[str, int] = {}
        observed_order = []
        for row in range(2, ws.max_row + 1):
            asin = normalize_asin(ws.cell(row, asin_col).value)
            if asin:
                row_by_asin[asin] = row
                observed_order.append(asin)
        expected_order = [normalize_asin(rec.get("asin")) for rec in ordered]
        if observed_order != expected_order:
            findings.append(_export_issue("", "workbook_row_order", sheet_name,
                                          EXPORT_VALUE_MISMATCH,
                                          "工作表 ASIN 顺序或集合与产品记录不一致",
                                          expected_order, observed_order))
        image_rows = _workbook_images_by_row(ws) if image_sheet else set()
        for seq, rec in enumerate(ordered, 1):
            asin = normalize_asin(rec.get("asin"))
            row = row_by_asin.get(asin)
            if row is None:
                findings.append(_export_issue(asin, "workbook_row", sheet_name, EXPORT_MISSING,
                                              "导出工作表缺少该 ASIN 的数据行"))
                continue
            expected_values = values_fn(rec, seq)
            for col, expected in enumerate(expected_values, 1):
                if image_sheet and col == 1:
                    continue
                actual = ws.cell(row, col).value
                if not _same_display_value(expected, actual):
                    display_column = headers[col - 1]
                    findings.append(_export_issue(
                        asin, "export.%s" % display_column, display_column,
                        EXPORT_VALUE_MISMATCH, "Excel 单元格值与导出层预期不一致",
                        expected, actual))
            if image_sheet and rec.get("image_url") and row not in image_rows:
                findings.append(_export_issue(asin, "export.image", "图片", IMAGE_MISSING,
                                              "中文表存在图片链接，但该 ASIN 未嵌入图片"))
    return findings


def audit_field_closure(products: Iterable[Mapping], details: Optional[Iterable[Mapping]] = None,
                        rankings: Optional[Iterable[Mapping]] = None,
                        html_dir: Optional[str | Path | Iterable[str | Path]] = None,
                        run_dir: Optional[str | Path | Iterable[str | Path]] = None,
                        workbook_path: Optional[str | Path] = None,
                        translations: Optional[Mapping] = None) -> dict:
    """Audit products without mutating any input mapping."""
    products = list(products or [])
    ranking_html = _read_ranking_html(run_dir or html_dir)
    html_by_asin = _html_by_asin(html_dir)
    details_by = {normalize_asin(d.get("asin")): d for d in (details or []) if normalize_asin(d.get("asin"))}
    rankings_by = {normalize_asin(r.get("asin")): r for r in (rankings or []) if normalize_asin(r.get("asin"))}
    records: list[dict] = []
    for product in sorted(products, key=lambda p: normalize_asin(p.get("asin"))):
        asin = normalize_asin(product.get("asin"))
        detail = details_by.get(asin, {})
        ranking = rankings_by.get(asin, {})
        html = html_by_asin.get(asin) or ""
        # When a run directory is supplied, category evidence comes only from
        # ranking_*.html, while detail HTML remains available for other fields.
        if run_dir and ranking_html:
            html_for_category = ranking_html
        else:
            html_for_category = html
        for field in _FIELD_ORDER:
            records.append(_audit_one(field, asin, product, detail, ranking,
                                      html_for_category if field.startswith("category_") or field == "leaf_category" else html).to_dict())
        tr = (translations or {}).get(asin) if isinstance(translations, Mapping) else None
        tr = tr if isinstance(tr, Mapping) else {}
        for target, (source_key, column) in _TRANSLATION_FIELDS.items():
            source = product.get(source_key)
            if source in (None, ""):
                continue
            target_value = product.get(target) or tr.get(target)
            if _translation_residual(source, target_value):
                records.append(FieldClosureResult(
                    asin=asin, field=target, display_column=column,
                    source_status="present", raw_status="present",
                    canonical_status="present", derived_status="present" if target_value else "missing",
                    display_status="mismatch" if target_value else "missing",
                    classification=TRANSLATION_INCOMPLETE, severity="P1",
                    source_evidence=[source_key], raw_evidence=source,
                    canonical_value=source, derived_value=target_value,
                    display_value=target_value,
                    message="中文展示字段为空或仍保留西语原文/整句").to_dict())
    if workbook_path:
        records.extend(_audit_workbook(products, workbook_path, translations=translations))
    counts = {c: sum(1 for r in records if r["classification"] == c)
              for c in (SOURCE_MISSING, PARSER_MISSED, MAPPING_MISSED, DERIVED_MISSING,
                        TRANSLATION_INCOMPLETE, NOT_OBSERVED, EVIDENCE_UNAVAILABLE, EXPORT_MISSING,
                        EXPORT_VALUE_MISMATCH, IMAGE_MISSING)}
    counts["pass"] = sum(1 for r in records if r["classification"] == PASS)
    field_summary = {}
    for field in _FIELD_ORDER:
        rows = [r for r in records if r["field"] == field]
        field_summary[field] = {c: sum(1 for r in rows if r["classification"] == c)
                                for c in (PASS, SOURCE_MISSING, PARSER_MISSED, MAPPING_MISSED,
                                          DERIVED_MISSING, TRANSLATION_INCOMPLETE, NOT_OBSERVED, EVIDENCE_UNAVAILABLE,
                                          "ORIGINAL_PRICE_INVALID")}
    defect_classes = (SOURCE_MISSING, PARSER_MISSED, MAPPING_MISSED, DERIVED_MISSING,
                      "ORIGINAL_PRICE_INVALID", EXPORT_MISSING, EXPORT_VALUE_MISMATCH,
                      IMAGE_MISSING, TRANSLATION_INCOMPLETE)
    return {
        "summary": {"total_skus": len(products), "fields_checked": len(_FIELD_ORDER), **counts},
        "coverage_summary": {
            "source_present": sum(1 for r in records if r["source_status"] == "present"),
            "not_observed": counts[NOT_OBSERVED],
            "evidence_unavailable": counts[EVIDENCE_UNAVAILABLE],
        },
        "defect_summary": {c: sum(1 for r in records if r["classification"] == c)
                           for c in defect_classes},
        "field_summary": field_summary,
        "records": records,
    }


def render_markdown(report: Mapping) -> str:
    summary = report.get("summary", {})
    lines = ["# Field Closure Audit", "", "## Summary", "",
             f"SKU: {summary.get('total_skus', 0)}", "",
             f"SOURCE_MISSING: {summary.get(SOURCE_MISSING, 0)}",
             f"PARSER_MISSED: {summary.get(PARSER_MISSED, 0)}",
             f"MAPPING_MISSED: {summary.get(MAPPING_MISSED, 0)}",
             f"DERIVED_MISSING: {summary.get(DERIVED_MISSING, 0)}",
             f"TRANSLATION_INCOMPLETE: {summary.get(TRANSLATION_INCOMPLETE, 0)}",
             f"NOT_OBSERVED: {summary.get(NOT_OBSERVED, 0)}",
             f"EVIDENCE_UNAVAILABLE: {summary.get(EVIDENCE_UNAVAILABLE, 0)}",
             f"PASS: {summary.get('pass', 0)}", "", "## By Field", "",
             "| Field | PASS | SOURCE_MISSING | PARSER_MISSED | MAPPING_MISSED | DERIVED_MISSING | NOT_OBSERVED | EVIDENCE_UNAVAILABLE |",
             "|---|---:|---:|---:|---:|---:|---:|---:|"]
    for field, counts in (report.get("field_summary") or {}).items():
        lines.append("| %s | %d | %d | %d | %d | %d | %d | %d |" % (
            field, counts.get(PASS, 0), counts.get(SOURCE_MISSING, 0),
            counts.get(PARSER_MISSED, 0), counts.get(MAPPING_MISSED, 0),
            counts.get(DERIVED_MISSING, 0), counts.get(NOT_OBSERVED, 0),
            counts.get(EVIDENCE_UNAVAILABLE, 0)))
    coverage = report.get("coverage_summary") or {}
    lines += ["", "## Coverage", "",
              "- Source present: %d" % coverage.get("source_present", 0),
              "- Not observed on saved page: %d" % coverage.get("not_observed", 0),
              "- Saved page unavailable: %d" % coverage.get("evidence_unavailable", 0),
              "", "## Defects", ""]
    issues = [r for r in report.get("records", []) if r.get("severity") != "INFO"]
    if not issues:
        lines.append("无闭环问题。")
    else:
        for r in issues:
            lines += [f"### {r.get('asin')}", "", f"**{r.get('field')}**（{r.get('display_column')}）",
                      f"- Raw: `{r.get('raw_evidence')}`",
                      f"- Canonical: `{r.get('canonical_value')}`",
                      f"- Excel: `{r.get('display_value')}`",
                      f"- Classification: **{r.get('classification')}** ({r.get('severity')})",
                      f"- {r.get('message')}", ""]
    return "\n".join(lines) + "\n"


def write_report(report: Mapping, out: str | Path, md_out: Optional[str | Path] = None) -> None:
    path = Path(out)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    md_path = Path(md_out) if md_out else path.with_suffix(".md")
    md_path.parent.mkdir(parents=True, exist_ok=True)
    md_path.write_text(render_markdown(report), encoding="utf-8")
