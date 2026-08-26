# -*- coding: utf-8 -*-
"""
第二轮结构整改 V2 —— 工作簿构建
输入：
  _v2_selected.json            —— V2 数据（预处理脚本产出）
  当前主文件 选品优化版.xlsx     —— 提取 150 张图片 + 类目规划 + 排行榜URL
  _translations.json           —— 商品名称中文翻译（按 row_idx）
输出：
  AmazonES_产品清单与提取信息_结构优化V2.xlsx （5 张表，不覆盖任何现有文件）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
import json
import os

SRC = r"E:\amazon_es\.worktrees\reconnaissance\AmazonESBestseller\outputs\amazon_es_catalog_20260825"
DATA = os.path.join(SRC, "_v2_selected.json")
PREV = os.path.join(SRC, "AmazonES_产品清单与提取信息_选品优化版.xlsx")
TRANSL = os.path.join(SRC, "_translations.json")
OUT = os.path.join(SRC, "AmazonES_产品清单与提取信息_结构优化V2.xlsx")

records = json.load(open(DATA, encoding='utf-8'))
transl = json.load(open(TRANSL, encoding='utf-8'))
print("records:", len(records), "| translations keys:", len(transl))

# 读取前版工作簿：图片 + 类目规划 + 排行榜URL
prev = openpyxl.load_workbook(PREV, data_only=True)
prev_img = openpyxl.load_workbook(PREV, data_only=False)  # 图片需要非 data_only
ws_prev_main = prev_img['选品清单']
images = []
for im in ws_prev_main._images:
    data = im.ref.getvalue()
    row0 = im.anchor._from.row  # 0-based
    images.append((row0, BytesIO(data), im.width, im.height))
print("提取图片:", len(images))

# 类目规划：原样迁移（38 类目）
cat_rows = list(prev['类目规划'].iter_rows(values_only=True))
cat_header = [str(h).strip() if h else '' for h in cat_rows[0]]
cat_data = [r for r in cat_rows[1:] if r and any(c is not None and str(c).strip() != '' for c in r)]
print("类目规划:", len(cat_data), "行")

# 排行榜URL（按行序映射到记录）
ws_rank_prev = prev['排行榜记录']
rank_urls = []
for r in ws_rank_prev.iter_rows(min_row=2, values_only=True):
    rank_urls.append(str(r[10]).strip() if len(r) > 10 and r[10] else '')

# ---------- 样式 ----------
HDR_FILL = PatternFill('solid', fgColor='1F4E78')
HDR_FONT = Font(bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(bold=True, size=14, color='1F4E78')
STAT_LABEL = Font(bold=True, size=10)
THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(wrap_text=True, vertical='top')

def put(ws, r, c, v):
    ws.cell(row=r, column=c, value=v)

def to_num(v):
    if v is None or str(v).strip() == '': return None
    try:
        f = float(str(v).replace(',', '.').replace('€', '').strip())
        return f
    except ValueError:
        return None

def stats_area(ws, ncols):
    n = len(records)
    l1s = {r['采集类目中文'] for r in records if r['采集类目中文']}
    leafs = {r['bsr_leaf_cat'] for r in records if r['bsr_leaf_cat']}
    n_bought = sum(1 for r in records if r['monthly_bought_raw'])
    n_disc = sum(1 for r in records if r['discount_rate'] != '')
    stats = [('商品数', n), ('一级类目数', len(l1s)), ('细分类目数', len(leafs)),
             ('有月购买量商品数', n_bought), ('有折扣商品数', n_disc), ('数据采集时间', '2026-08-25')]
    c = 1
    for label, val in stats:
        put(ws, 2, c, label); ws.cell(row=2, column=c).font = STAT_LABEL
        put(ws, 2, c + 1, val); ws.cell(row=2, column=c + 1).font = Font(size=10, bold=True)
        c += 2
    put(ws, 1, 1, 'Amazon.es 畅销商品内部选品数据库 · 选品清单（V2）')
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)

# ---------- Sheet1 选品清单 ----------
wb = openpyxl.Workbook()
ws = wb.active
ws.title = '选品清单'
HEAD1 = ['图片','序号','商品名称','当前价格','原价','折扣率','一级类目','二级类目','三级类目',
         '细分类目','月购买量','类目排名','规格','商品详情摘要','上架时间','商品链接','图片链接',
         'ASIN','选品状态','研究备注']
WIDTH1 = [10,5,42,9,9,8,11,8,8,20,10,9,28,48,11,26,34,12,10,16]
for i, h in enumerate(HEAD1, 1):
    cell = ws.cell(row=3, column=i, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = WIDTH1[i-1]
stats_area(ws, len(HEAD1))

# 序号：按一级类目分组重排
seq = {}
idx = 0
for r in records:
    l1 = r['采集类目中文']
    seq.setdefault(l1, 0); seq[l1] += 1
    r['_seq'] = seq[l1]

for i, r in enumerate(records):
    row = 4 + i
    cur_n = to_num(r['current_price']); orig_n = to_num(r['original_price'])
    disc_n = None
    if r['discount_rate'] != '':
        try: disc_n = float(str(r['discount_rate']).replace(',', '.'))
        except ValueError: pass
    vals = [None, r['_seq'], r['title_es_raw'], cur_n, orig_n, disc_n,
            r['采集类目中文'], '', '', r['bsr_leaf_cat'],
            r['monthly_bought_min'] or '', r['bsr_leaf_rank'], r['spec_v2'], r['summary_v2'],
            r['date_first_available'], r['product_url'], r['image_url'], r['asin'],
            r['选品状态'], r['研究备注']]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.border = BORDER
        if c in (4, 5) and isinstance(v, (int, float)):
            cell.number_format = '#,##0.00'
        if c == 6 and isinstance(v, (int, float)):
            cell.number_format = '0%'
        if c in (3, 13, 14, 10):
            cell.alignment = WRAP
        if c in (16, 17):
            cell.font = Font(color='0563C1', underline='single', size=9)
        if c == 2:
            cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 78

# 图片（锚定行不变）
img_map = {}
for (row0, bio, w, h) in images:
    img_map[row0] = (bio, w, h)
for row0, (bio, w, h) in img_map.items():
    img = XLImage(bio)
    img.width = (w if w else 70)
    img.height = (h if h else 70)
    ws.add_image(img, 'A%d' % (row0 + 1))
print("选品清单图片:", len(img_map))

# 下拉 + 冻结 + 筛选
dv = DataValidation(type='list', formula1='"待评估,重点关注,暂不考虑,已研究"', allow_blank=True)
ws.add_data_validation(dv)
dv.add('S4:S%d' % (3 + len(records)))
ws.freeze_panes = 'A4'
ws.auto_filter.ref = 'A3:T%d' % (3 + len(records))

# ---------- Sheet2 排行榜记录（每榜单一行） ----------
ws2 = wb.create_sheet('排行榜记录')
HEAD2 = ['index','asin','category_l1','category_l2','category_l3','leaf_category','browse_node_id',
         'category_rank','monthly_bought_raw','monthly_bought_min','ranking_source_url','collected_at']
for i, h in enumerate(HEAD2, 1):
    cell = ws2.cell(row=1, column=i, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    ws2.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16
row2 = 2
for i, r in enumerate(records):
    segs = r['bsr_segments']
    if not segs:
        continue
    url = rank_urls[i] if i < len(rank_urls) else ''
    for cat, rank in segs:
        # 每行 = 一个真实榜单条目；leaf_category 记录该榜单名（主榜行为顶级榜单名，细分行为细分类目名）
        vals = [row2 - 1, r['asin'], r['采集类目中文'], '', '', cat, '',
                rank, r['monthly_bought_raw'] or '', r['monthly_bought_min'] or '', url, r['first_seen'][:19]]
        for c, v in enumerate(vals, 1):
            cell = ws2.cell(row=row2, column=c, value=v)
            cell.border = BORDER
        row2 += 1
print("排行榜记录行数:", row2 - 2)
ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = 'A1:L%d' % (row2 - 1)

# ---------- Sheet3 后台数据 ----------
ws3 = wb.create_sheet('后台数据')
HEAD3 = ['asin','parent_asin','parent_asin_status','title_es_raw','brand_raw','brand','price_legacy',
         'current_price','original_price','currency','discount_rate','rating','review_count',
         'monthly_bought_raw','monthly_bought_min','image_url','product_url','details_json','details_raw',
         'specification_legacy','date_first_available_raw','date_first_available','first_seen','last_seen',
         'ranking_count','best_rank_legacy','bsr_segments','bsr_main_cat','bsr_main_rank','bsr_leaf_cat',
         'bsr_leaf_rank','上榜细分类目数','upc','gtin','image_path','image_download_status',
         'image_download_error','详情状态','图片状态','采集类目中文','采集类目西语','选品状态','研究备注','重采标志']
for i, h in enumerate(HEAD3, 1):
    cell = ws3.cell(row=1, column=i, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = BORDER
    ws3.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 16
for i, r in enumerate(records):
    row = 2 + i
    bsr_json = json.dumps(r['bsr_segments'], ensure_ascii=False)
    vals = [r['asin'], r['parent_asin'], r['parent_asin_status'], r['title_es_raw'], r['brand_raw'],
            r['brand'], r['price_legacy'], to_num(r['current_price']), to_num(r['original_price']),
            r['currency'], r['discount_rate'], r['rating'], r['review_count'],
            r['monthly_bought_raw'], r['monthly_bought_min'], r['image_url'], r['product_url'],
            r['details_json'], r['details_raw'], r['specification_legacy'],
            r['date_first_available_raw'], r['date_first_available'], r['first_seen'], r['last_seen'],
            r['ranking_count'], r['best_rank_legacy'], bsr_json, r['bsr_main_cat'], r['bsr_main_rank'],
            r['bsr_leaf_cat'], r['bsr_leaf_rank'], r['上榜细分类目数'], r['upc'], r['gtin'],
            r['image_path'], r['image_download_status'], r['image_download_error'],
            r['详情状态'], r['图片状态'], r['采集类目中文'], r['采集类目西语'],
            r['选品状态'], r['研究备注'], r['重采标志']]
    for c, v in enumerate(vals, 1):
        cell = ws3.cell(row=row, column=c, value=v)
        cell.border = BORDER
        if c in (8, 9) and isinstance(v, (int, float)):
            cell.number_format = '#,##0.00'
ws3.freeze_panes = 'B2'
ws3.auto_filter.ref = 'A1:%s%d' % (openpyxl.utils.get_column_letter(len(HEAD3)), len(records) + 1)

# ---------- Sheet4 类目规划 ----------
ws4 = wb.create_sheet('类目规划')
for i, h in enumerate(cat_header, 1):
    cell = ws4.cell(row=1, column=i, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    ws4.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20
for i, r in enumerate(cat_data):
    for c, v in enumerate(r, 1):
        if v is None: continue
        cell = ws4.cell(row=2 + i, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP
print("类目规划写入:", len(cat_data))

# ---------- Sheet5 商品名称中文对照（小表，无图） ----------
ws5 = wb.create_sheet('商品名称中文对照')
HEAD5 = ['序号','ASIN','商品名称(西语)','商品名称(中文)']
for i, h in enumerate(HEAD5, 1):
    cell = ws5.cell(row=1, column=i, value=h)
    cell.font = HDR_FONT; cell.fill = HDR_FILL
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = BORDER
    ws5.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 18
for i, r in enumerate(records):
    t = transl.get(str(i), {})
    title_zh = t.get('title_zh', '') if isinstance(t, dict) else ''
    vals = [r['_seq'], r['asin'], r['title_es_raw'], title_zh]
    for c, v in enumerate(vals, 1):
        cell = ws5.cell(row=2 + i, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP
    ws5.row_dimensions[2 + i].height = 30
zh_filled = sum(1 for i, r in enumerate(records)
                if isinstance(transl.get(str(i), {}), dict) and transl.get(str(i), {}).get('title_zh'))
print("中文名填充:", zh_filled, "/", len(records))
ws5.freeze_panes = 'A2'
ws5.auto_filter.ref = 'A1:D%d' % (len(records) + 1)

wb.save(OUT)
print("已保存:", OUT)
