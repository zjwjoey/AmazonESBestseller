# -*- coding: utf-8 -*-
"""Read-only audit of the round-1 output workbook (选品优化版) to ground the review."""
import openpyxl, os, hashlib

P = r"E:\amazon_es\.worktrees\reconnaissance\AmazonESBestseller\outputs\amazon_es_catalog_20260825\AmazonES_产品清单与提取信息_选品优化版.xlsx"
print("exists:", os.path.exists(P), "| size:", os.path.getsize(P))
print("md5:", hashlib.md5(open(P,'rb').read()).hexdigest())

wb = openpyxl.load_workbook(P, read_only=True, data_only=True)
print("sheets:", wb.sheetnames)

def sheet_stats(name, header_row_idx, data_start):
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    print("\n===== %s | dims: %d rows x %d cols =====" % (name, ws.max_row, ws.max_column))
    hdr = rows[header_row_idx-1]
    hdr = [str(h).strip() if h is not None else '' for h in hdr]
    print("HEADER (%d): %s" % (len(hdr), hdr))
    data = rows[data_start-1:]
    # fill rate per column over data rows
    print("fill rate per col (n=%d):" % len(data))
    for i, h in enumerate(hdr):
        if not h:
            continue
        c = sum(1 for r in data if i < len(r) and r[i] is not None and str(r[i]).strip() != '')
        print("  [%2d] %-22s %3d/%d" % (i, h, c, len(data)))
    return hdr, data

print("\n############ TOP-ROW AREA of 选品清单 (stats block) ############")
ws = wb['选品清单']
for r in ws.iter_rows(min_row=1, max_row=5, values_only=True):
    print(r)

hdr, data = sheet_stats('选品清单', 5, 6)
# sample 2 data rows (truncate long cells)
for r in data[:2]:
    out = []
    for i, v in enumerate(r):
        if v is None:
            out.append('-')
        else:
            s = str(v).replace('\n', ' ')
            out.append(s[:26] + ('…' if len(s) > 26 else ''))
    print("sample:", out)

# dropdowns / images / freeze
wb2 = openpyxl.load_workbook(P, data_only=True)
ws2 = wb2['选品清单']
print("\n选品清单 freeze_panes:", ws2.freeze_panes, "| auto_filter:", ws2.auto_filter.ref)
print("images:", len(ws2._images))
print("data_validations:", len(ws2.data_validations.dataValidation))
for dv in ws2.data_validations.dataValidation:
    print("  dv type:", dv.type, "formula1:", dv.formula1, "range:", dv.sqref)

# 排行榜记录
hdr2, data2 = sheet_stats('排行榜记录', 1, 2)
asins = [r[1] for r in data2 if len(r) > 1 and r[1]]
from collections import Counter
cc = Counter(asins)
dups = {k: v for k, v in cc.items() if v > 1}
print("dup ASIN rows:", len(dups), "| total rows:", len(data2))

# 后台数据
hdr3, data3 = sheet_stats('后台数据', 1, 2)
keycols = ['asin','parent_asin','brand','brand_raw','current_price','original_price','currency',
           'rating','review_count','image_url','product_url','details_json','details_raw',
           'specification','specification_legacy','date_first_available','date_first_available_raw',
           'first_seen','last_seen','monthly_bought_text','monthly_bought_raw','best_rank','best_rank_legacy']
for k in keycols:
    if k in hdr3:
        i = hdr3.index(k)
        c = sum(1 for r in data3 if i < len(r) and r[i] is not None and str(r[i]).strip() != '')
        print("  [后台] %-30s %3d/%d" % (k, c, len(data3)))

# 类目规划
sheet_stats('类目规划', 1, 2)

# 选品清单中文对照
hdr5, data5 = sheet_stats('选品清单中文对照', 1, 2)
ws5 = wb2['选品清单中文对照']
print("images:", len(ws5._images))
print("freeze:", ws5.freeze_panes)
