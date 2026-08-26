# -*- coding: utf-8 -*-
"""cli.py 测试：统一入口（enrich/qa/export 离线主链 + collect 联网拒离线）。

不联网：只测离线子命令与参数纪律；collect 的真实采集不在此覆盖。
"""
import json
import base64
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

from amazon_es_bestseller.cli import main

REPO = Path(__file__).resolve().parent.parent


def test_cli_help_lists_subcommands(capsys):
    with pytest.raises(SystemExit) as ei:
        main(["--help"])
    assert ei.value.code == 0
    out = capsys.readouterr().out
    for cmd in ("collect", "enrich", "qa", "audit-fields", "export", "select-quota", "translate-ds"):
        assert cmd in out


def test_cli_select_quota_writes_grouped_manifest(tmp_path):
    rankings = tmp_path / "rankings.json"
    config = tmp_path / "config.json"
    out = tmp_path / "manifest.json"
    rankings.write_text(json.dumps([
        {"asin": "h1", "ranking_source_url": "https://www.amazon.es/gp/bestsellers/kitchen/"},
        {"asin": "d1", "ranking_source_url": "https://www.amazon.es/gp/bestsellers/diy/"},
    ]), encoding="utf-8")
    config.write_text(json.dumps([
        {"group": "hogar", "url": "https://www.amazon.es/gp/bestsellers/kitchen/", "quota": 1},
        {"group": "diy", "url": "https://www.amazon.es/gp/bestsellers/diy/", "quota": 1},
    ]), encoding="utf-8")
    assert main(["select-quota", "--rankings", str(rankings), "--config", str(config), "--out", str(out)]) == 0
    manifest = json.loads(out.read_text(encoding="utf-8"))
    assert manifest["summary"] == {"hogar": 1, "diy": 1, "total": 2}
    assert [r["asin"] for r in manifest["records"]] == ["H1", "D1"]


def test_cli_translate_ds_writes_asin_map(tmp_path, monkeypatch):
    products = tmp_path / "products.json"
    out = tmp_path / "translations.json"
    products.write_text(json.dumps([{"asin": "b1", "title_es_raw": "Taladro"}], ensure_ascii=False), encoding="utf-8")

    class FakeTranslator:
        def __init__(self, **kwargs):
            pass

        def translate_record(self, record):
            return {"asin": record["asin"].upper(), "title_zh": "电钻", "translation_status": "success"}

        def save_cache(self):
            pass

    import amazon_es_bestseller.translation.ds as ds
    monkeypatch.setattr(ds, "DeepSeekTranslator", FakeTranslator)
    assert main(["translate-ds", "--products", str(products), "--out", str(out)]) == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert saved["B1"]["title_zh"] == "电钻"


def test_cli_audit_fields_writes_json_and_markdown(tmp_path, capsys):
    products = tmp_path / "products.json"
    details = tmp_path / "details.json"
    rankings = tmp_path / "rankings.json"
    out = tmp_path / "field_closure.json"
    products.write_text(json.dumps([{"asin": "B000000001", "title_es_raw": "Caja",
                                     "title_zh": "", "product_url": "https://www.amazon.es/dp/B000000001",
                                     "image_url": "https://img"}], ensure_ascii=False), encoding="utf-8")
    details.write_text("[]", encoding="utf-8")
    rankings.write_text("[]", encoding="utf-8")
    assert main(["audit-fields", "--products", str(products), "--details", str(details),
                 "--rankings", str(rankings), "--out", str(out)]) == 0
    report = json.loads(out.read_text(encoding="utf-8"))
    assert report["summary"]["total_skus"] == 1
    assert (tmp_path / "field_closure.md").exists()


def test_collect_rejects_offline(capsys):
    # --offline 是全局参数，置于子命令前
    with pytest.raises(SystemExit) as ei:
        main(["--offline", "collect", "--urls", "https://www.amazon.es/zgbs/1"])
    assert ei.value.code == 2
    assert "联网" in capsys.readouterr().err


def test_collect_requires_urls(capsys):
    with pytest.raises(SystemExit) as ei:
        main(["collect"])
    assert ei.value.code == 2


def _real_data():
    p = REPO / "product_details.json"
    return p if p.exists() else None


def test_enrich_qa_export_offline_chain(tmp_path, capsys):
    """30 条真实数据走 cli enrich → qa → export（0 P0/P1）。"""
    src = _real_data()
    if src is None:
        pytest.skip("product_details.json 不在仓库")
    prod_out = tmp_path / "products.json"
    qa_out = tmp_path / "qa.json"
    xlsx_out = tmp_path / "选品清单.xlsx"

    assert main(["enrich", "--legacy", str(src), "--out", str(prod_out)]) == 0
    products = json.loads(prod_out.read_text(encoding="utf-8"))
    assert len(products) == 30
    # 构造型 BSR 已丢弃：无任何 detail_bsr_segments
    assert all(not p.get("detail_bsr_segments") for p in products)

    assert main(["qa", "--products", str(prod_out), "--out", str(qa_out)]) == 0
    qa = json.loads(qa_out.read_text(encoding="utf-8"))
    assert qa["summary"]["total_products"] == 30
    assert qa["summary"]["fail_count"] == 0
    assert qa["summary"]["source_conflict_count"] == 0
    # 0 P0/P1（P2 缺失类 WARN 允许）
    p0p1 = [i for r in qa["records"] for i in r["issues"] if i["severity"] in ("P0", "P1")]
    assert not p0p1

    assert main(["export", "--products", str(prod_out), "--out", str(xlsx_out)]) == 0
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_out)
    # B3x 新契约：3 张表
    assert wb.sheetnames == ["类目规划", "西班牙语选品清单", "中文选品清单"]


def test_enrich_missing_input_file(tmp_path):
    with pytest.raises(SystemExit) as ei:
        main(["enrich", "--rankings", str(tmp_path / "nope.json"),
              "--out", str(tmp_path / "x.json")])
    assert "找不到输入文件" in str(ei.value.code)


def _blocked_products():
    """含 1 条 P0（ASIN 非法）+ 1 条合法记录的商品表（测试 QA 导出门禁）。"""
    return [
        {"asin": "B078C6QR1C", "title_es_raw": "Protector",
         "current_price_raw": "12,62 €"},
        {"asin": None, "title_es_raw": "sin asin", "current_price_raw": "1,00 €"},
    ]


def test_export_qa_gate_blocks_on_p0(tmp_path, capsys):
    """QA_RULES §31：存在 P0/P1 → 默认拒绝导出（保留上游错误证据）。"""
    prod_out = tmp_path / "products.json"
    prod_out.write_text(
        json.dumps(_blocked_products()), encoding="utf-8")
    with pytest.raises(SystemExit) as ei:
        main(["export", "--products", str(prod_out),
              "--out", str(tmp_path / "out.xlsx")])
    msg = str(ei.value.code)
    assert "拒绝导出" in msg
    assert "ASIN_INVALID" in msg
    assert not (tmp_path / "out.xlsx").exists()     # 未产出 Excel


def test_export_qa_gate_force_bypasses(tmp_path):
    """--force：跳过 QA 门禁强制导出（明确授权，不静默）。"""
    prod_out = tmp_path / "products.json"
    prod_out.write_text(
        json.dumps(_blocked_products()), encoding="utf-8")
    xlsx_out = tmp_path / "out.xlsx"
    assert main(["export", "--products", str(prod_out),
                 "--out", str(xlsx_out), "--force"]) == 0
    assert xlsx_out.exists()


def test_export_accepts_images_and_category_planning(tmp_path):
    products = tmp_path / "products.json"
    products.write_text(json.dumps([{
        "asin": "B078C6QR1C",
        "product_url": "https://www.amazon.es/dp/B078C6QR1C",
        "image_url": "https://m.media-amazon.com/images/I/x.jpg",
        "current_price": 1,
    }]), encoding="utf-8")
    images = tmp_path / "images"
    images.mkdir()
    png = base64.b64decode(
        "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg==")
    (images / "B078C6QR1C.png").write_bytes(png)
    planning = tmp_path / "planning.json"
    planning.write_text(json.dumps([{"#": 1, "中文一级类目": "家居与厨房"}], ensure_ascii=False), encoding="utf-8")
    out = tmp_path / "out.xlsx"
    assert main(["export", "--products", str(products), "--images-dir", str(images),
                  "--category-planning", str(planning), "--out", str(out), "--force"]) == 0
    wb = load_workbook(out)
    assert wb["类目规划"].max_row == 2
    assert len(wb["中文选品清单"]._images) == 1
    assert len(wb["西班牙语选品清单"]._images) == 0
