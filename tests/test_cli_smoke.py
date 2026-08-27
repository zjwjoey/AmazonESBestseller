# -*- coding: utf-8 -*-
"""CLI smoke/vertical tests: all paths stay offline and use tiny fixtures."""
import json
from pathlib import Path

import pytest

from amazon_es_bestseller.cli import main


def test_reparse_details_prints_its_own_report_and_supports_multiple_dirs(tmp_path, capsys):
    state = tmp_path / "state.json"
    state.write_text("{}", encoding="utf-8")
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()
    (first / "B000000001.html").write_text(
        '<input id="ASIN" value="B000000001"><div id="productTitle">Caja</div>',
        encoding="utf-8",
    )
    (second / "B000000002.html").write_text(
        '<input id="ASIN" value="B000000002"><div id="productTitle">Otra caja</div>',
        encoding="utf-8",
    )
    out = tmp_path / "details.json"

    assert main(["reparse-details", "--html-dir", str(first), str(second),
                 "--state", str(state), "--out", str(out)]) == 0
    captured = capsys.readouterr()
    assert "reparse-details" in captured.out
    assert "repair-cache" not in captured.out
    assert "B000000001" in out.read_text(encoding="utf-8")
    saved = json.loads(out.read_text(encoding="utf-8"))
    saved_by_asin = {row["asin"]: row for row in saved}
    assert set(saved_by_asin) == {"B000000001", "B000000002"}
    assert saved_by_asin["B000000001"]["detail_schema_version"]
    persisted = json.loads(state.read_text(encoding="utf-8"))
    assert persisted["B000000002"]["detail_schema_version"] == saved_by_asin["B000000002"]["detail_schema_version"]


def test_reparse_details_first_directory_wins_duplicate_asin(tmp_path):
    state = tmp_path / "state.json"
    state.write_text("{}", encoding="utf-8")
    first = tmp_path / "first"
    second = tmp_path / "second"
    first.mkdir()
    second.mkdir()
    (first / "B000000001.html").write_text(
        '<input id="ASIN" value="B000000001"><div id="productTitle">优先版本</div>',
        encoding="utf-8",
    )
    (second / "B000000001.html").write_text(
        '<input id="ASIN" value="B000000001"><div id="productTitle">重复版本</div>',
        encoding="utf-8",
    )
    out = tmp_path / "details.json"
    assert main(["reparse-details", "--html-dir", str(first), str(second),
                 "--state", str(state), "--out", str(out)]) == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    saved_by_asin = {row["asin"]: row for row in saved}
    assert saved_by_asin["B000000001"]["title_es_raw"] == "优先版本"


def test_collect_rankings_only_smoke_uses_fake_browser(tmp_path, monkeypatch):
    from amazon_es_bestseller.access import browser
    from amazon_es_bestseller.collection import ranking

    class FakeSession:
        def __init__(self, **kwargs):
            self.headless = kwargs["headless"]

        def __enter__(self):
            return self

        def __exit__(self, exc_type, exc, tb):
            return False

    calls = []

    def fake_collect_rankings(urls, session, out_dir):
        calls.append((urls, session.headless, out_dir))
        return [{"asin": "B000000001", "ranking_source_url": urls[0]}]

    monkeypatch.setattr(browser, "BrowserSession", FakeSession)
    monkeypatch.setattr(ranking, "collect_rankings", fake_collect_rankings)
    out_dir = tmp_path / "run"
    assert main(["collect", "--rankings-only", "--urls", "https://example.invalid/rank",
                 "--out-dir", str(out_dir)]) == 0
    assert calls and calls[0][0] == ["https://example.invalid/rank"]
    assert json.loads((out_dir / "rankings.json").read_text(encoding="utf-8"))[0]["asin"] == "B000000001"


def test_finalize_manifest_running_defaults_to_success():
    from amazon_es_bestseller.run_manifest import create_manifest, finalize_manifest, update_manifest

    manifest = update_manifest(create_manifest("run-001", started_at="x"), status="running")
    finalized = finalize_manifest(manifest)
    assert finalized["status"] == "success"
    assert finalized["finished_at"]


def test_run_manifest_round_trip_and_stage_updates(tmp_path):
    from amazon_es_bestseller.run_manifest import (
        create_manifest, finalize_manifest, load_manifest, update_manifest,
        write_manifest,
    )

    manifest = create_manifest("run-001", started_at="2026-08-27T00:00:00+08:00",
                              git_commit="abc123", config_hash="cfg")
    manifest = update_manifest(manifest, status="running", ranking_records=5,
                               unique_asins=4, detail_planned=4)
    manifest = finalize_manifest(manifest, status="success", export_status="success",
                                 final_workbook="out.xlsx")
    path = write_manifest(manifest, tmp_path / "run.json")
    loaded = load_manifest(path)
    assert loaded["run_id"] == "run-001"
    assert loaded["ranking_records"] == 5
    assert loaded["detail_planned"] == 4
    assert loaded["status"] == "success"
    assert loaded["final_workbook"] == "out.xlsx"


def test_select_quota_global_uniqueness_and_shortfall(tmp_path):
    rankings = tmp_path / "rankings.json"
    config = tmp_path / "config.json"
    out = tmp_path / "manifest.json"
    rankings.write_text(json.dumps([
        {"asin": "A000000001", "category_group": "hogar"},
        {"asin": "A000000002", "category_group": "hogar"},
        {"asin": "A000000002", "category_group": "diy"},
        {"asin": "A000000003", "category_group": "diy"},
    ]), encoding="utf-8")
    config.write_text(json.dumps([
        {"group": "hogar", "quota": 2}, {"group": "diy", "quota": 1},
    ]), encoding="utf-8")
    assert main(["select-quota", "--rankings", str(rankings), "--config", str(config),
                 "--out", str(out)]) == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert saved["summary"]["total"] == 3
    assert len({r["asin"] for r in saved["records"]}) == 3

    config.write_text(json.dumps([
        {"group": "hogar", "quota": 2}, {"group": "diy", "quota": 2},
    ]), encoding="utf-8")
    with pytest.raises(SystemExit) as ei:
        main(["select-quota", "--rankings", str(rankings), "--config", str(config),
              "--out", str(out)])
    assert "QUOTA_UNIQUE_SHORTFALL" in str(ei.value.code)


def test_translate_ds_isolates_partial_and_failed_records(tmp_path, monkeypatch, capsys):
    products = tmp_path / "products.json"
    out = tmp_path / "translations.json"
    products.write_text(json.dumps([
        {"asin": "B000000001", "title_es_raw": "Uno"},
        {"asin": "B000000002", "title_es_raw": "Dos"},
        {"asin": "B000000003", "title_es_raw": "Tres"},
    ]), encoding="utf-8")
    saves = []

    class FakeTranslator:
        def __init__(self, **kwargs):
            self.cache_path = kwargs.get("cache_path")

        def translate_record(self, record):
            status = {"B000000001": "success", "B000000002": "partial",
                      "B000000003": "failed"}[record["asin"]]
            return {"asin": record["asin"], "translation_status": status,
                    "title_zh": record["asin"] if status != "failed" else ""}

        def save_cache(self):
            saves.append(1)

    import amazon_es_bestseller.translation.ds as ds
    monkeypatch.setattr(ds, "DeepSeekTranslator", FakeTranslator)
    monkeypatch.setattr("builtins.input", lambda prompt: "YES")
    assert main(["translate-ds", "--products", str(products), "--out", str(out)]) == 0
    saved = json.loads(out.read_text(encoding="utf-8"))
    assert {v["translation_status"] for v in saved.values()} == {"success", "partial", "failed"}
    assert len(saves) == 3
    assert "成功 1、部分 1、失败 1" in capsys.readouterr().out


def test_translate_ds_offline_rejected_before_translator_creation(monkeypatch):
    import amazon_es_bestseller.translation.ds as ds

    class ExplodingTranslator:
        def __init__(self, **kwargs):
            raise AssertionError("translator must not be created in offline mode")

    monkeypatch.setattr(ds, "DeepSeekTranslator", ExplodingTranslator)
    with pytest.raises(SystemExit) as ei:
        main(["--offline", "translate-ds", "--products", "missing.json", "--out", "out.json"])
    assert "不能与 --offline 同用" in str(ei.value)


def test_translate_ds_requires_explicit_yes_before_first_request(tmp_path, monkeypatch):
    products = tmp_path / "products.json"
    products.write_text(json.dumps([{"asin": "B000000001", "title_es_raw": "Uno"}]), encoding="utf-8")
    calls = []

    class FakeTranslator:
        def __init__(self, **kwargs):
            calls.append("init")

        def translate_record(self, record):
            calls.append("translate")
            return {"asin": record["asin"], "translation_status": "success"}

        def save_cache(self):
            pass

    import amazon_es_bestseller.translation.ds as ds
    monkeypatch.setattr(ds, "DeepSeekTranslator", FakeTranslator)
    monkeypatch.setattr("builtins.input", lambda prompt: "NO")
    with pytest.raises(SystemExit) as ei:
        main(["translate-ds", "--products", str(products), "--out", str(tmp_path / "out.json")])
    assert "未确认" in str(ei.value)
    assert calls == []


def test_select_quota_to_export_minimal_offline_vertical_path(tmp_path, monkeypatch):
    rankings = tmp_path / "rankings.json"
    config = tmp_path / "config.json"
    quota_manifest = tmp_path / "quota.json"
    details = tmp_path / "details.json"
    translations = tmp_path / "translations.json"
    normalized = tmp_path / "normalized.json"
    products = tmp_path / "products.json"
    qa = tmp_path / "qa.json"
    closure = tmp_path / "closure.json"
    workbook = tmp_path / "out.xlsx"
    asins = ["B000000001", "B000000002"]
    rankings.write_text(json.dumps([{
        "asin": asins[0], "bestseller_rank": 1,
        "ranking_source_url": "https://www.amazon.es/Best-Sellers-Hogar/zgbs/1",
        "category_l1": "Hogar y cocina", "category_group": "hogar",
    }, {
        "asin": asins[1], "bestseller_rank": 1,
        "ranking_source_url": "https://www.amazon.es/Best-Sellers-DIY/zgbs/2",
        "category_l1": "Bricolaje y herramientas", "category_group": "diy",
    }]), encoding="utf-8")
    config.write_text(json.dumps([
        {"group": "hogar", "quota": 1}, {"group": "diy", "quota": 1},
    ]), encoding="utf-8")
    assert main(["select-quota", "--rankings", str(rankings), "--config", str(config),
                 "--out", str(quota_manifest)]) == 0
    selected = json.loads(quota_manifest.read_text(encoding="utf-8"))
    assert selected["summary"]["total"] == 2
    assert len({row["asin"] for row in selected["records"]}) == 2
    details.write_text(json.dumps([{
        "asin": asin, "title_es_raw": "Caja 2 piezas", "current_price_raw": "12,99 €",
        "brand_raw": "KRUPS", "selected_variation_raw": "Rojo",
        "attributes": [{"label_raw": "Número de piezas", "value_raw": "2"}],
        "feature_bullets_raw": ["Caja reutilizable"], "product_url": f"https://www.amazon.es/dp/{asin}",
        "image_url": "https://example.invalid/x.jpg",
    } for asin in asins]), encoding="utf-8")

    assert main(["enrich", "--rankings", str(rankings), "--details", str(details),
                 "--out", str(normalized)]) == 0
    normalized_records = json.loads(normalized.read_text(encoding="utf-8"))

    from amazon_es_bestseller.translation.ds import DeepSeekTranslator as RealTranslator

    class FakeTranslator:
        def __init__(self, **kwargs):
            pass

        @staticmethod
        def source_hash(record):
            return RealTranslator.source_hash(record)

        def translate_record(self, record):
            return {"asin": record["asin"], "translation_schema_version": 2,
                    "translation_source_hash": RealTranslator.source_hash(record),
                    "title_zh": "两件装收纳盒",
                    "selected_variation_zh": "红色", "specification_zh": "2件",
                    "product_details_zh": "品牌：Marca；数量：2",
                    "feature_bullets_zh": "可重复使用收纳盒",
                    "translation_status": "success"}

        def save_cache(self):
            pass

    import amazon_es_bestseller.translation.ds as ds
    monkeypatch.setattr(ds, "DeepSeekTranslator", FakeTranslator)
    monkeypatch.setattr("builtins.input", lambda prompt: "YES")
    assert main(["translate-ds", "--products", str(normalized), "--out", str(translations)]) == 0
    assert main(["enrich", "--rankings", str(rankings), "--details", str(details),
                 "--translations", str(translations), "--out", str(products)]) == 0
    assert main(["qa", "--products", str(products), "--out", str(qa)]) == 0
    assert main(["audit-fields", "--products", str(products), "--details", str(details),
                 "--rankings", str(rankings), "--translations", str(translations),
                 "--out", str(closure)]) == 0
    assert main(["export", "--products", str(products), "--details", str(details),
                 "--rankings", str(rankings), "--translations", str(translations),
                 "--out", str(workbook)]) == 0
    saved = json.loads(products.read_text(encoding="utf-8"))
    assert [row["asin"] for row in saved] == asins
    assert all(row["product_details_zh"] for row in saved)
    assert all(row["feature_bullets_zh"] for row in saved)
    assert json.loads(qa.read_text(encoding="utf-8"))["records"]
    assert json.loads(closure.read_text(encoding="utf-8"))["summary"]
    import openpyxl
    wb = openpyxl.load_workbook(workbook)
    assert wb.sheetnames == ["类目规划", "西班牙语选品清单", "中文选品清单"]
    assert wb["中文选品清单"].max_column == 26
    es_header = [cell.value for cell in wb["西班牙语选品清单"][1]]
    zh_header = [cell.value for cell in wb["中文选品清单"][1]]
    es_asins = [wb["西班牙语选品清单"].cell(row=i, column=es_header.index("ASIN") + 1).value
                for i in range(2, 4)]
    zh_asins = [wb["中文选品清单"].cell(row=i, column=zh_header.index("ASIN") + 1).value
                for i in range(2, 4)]
    assert es_asins == zh_asins == asins


def _export_gate_product(tmp_path):
    products = tmp_path / "products.json"
    products.write_text(json.dumps([{
        "asin": "B000000001", "title_es_raw": "Caja",
        "product_url": "https://www.amazon.es/dp/B000000001",
        "image_url": "https://example.invalid/x.jpg", "current_price": 1,
        "brand": "KRUPS",
    }]), encoding="utf-8")
    details = tmp_path / "details.json"
    details.write_text("[{\"asin\": \"B000000001\"}]", encoding="utf-8")
    return products, details


@pytest.mark.parametrize("classification", [
    "PARSER_MISSED", "MAPPING_MISSED", "DERIVED_MISSING", "TRANSLATION_INCOMPLETE",
])
def test_export_gate_blocks_closure_p1(tmp_path, monkeypatch, classification):
    products, details = _export_gate_product(tmp_path)
    import amazon_es_bestseller.qa.field_closure as closure
    monkeypatch.setattr(closure, "audit_field_closure", lambda *args, **kwargs: {
        "records": [{"asin": "B000000001", "classification": classification,
                     "severity": "P1", "message": "fixture"}]
    })
    with pytest.raises(SystemExit) as ei:
        main(["export", "--products", str(products), "--details", str(details),
              "--out", str(tmp_path / "blocked.xlsx")])
    assert "拒绝导出" in str(ei.value.code)


def test_export_gate_allows_source_missing_only(tmp_path, monkeypatch):
    products, details = _export_gate_product(tmp_path)
    import amazon_es_bestseller.qa.field_closure as closure
    monkeypatch.setattr(closure, "audit_field_closure", lambda *args, **kwargs: {
        "records": [{"asin": "B000000001", "classification": "SOURCE_MISSING",
                     "severity": "P2", "message": "fixture"}]
    })
    out = tmp_path / "allowed.xlsx"
    assert main(["export", "--products", str(products), "--details", str(details),
                 "--out", str(out)]) == 0
    assert out.exists()


@pytest.mark.parametrize("command", [
    "collect", "select-quota", "translate-ds", "enrich", "repair-cache",
    "reparse-details", "qa", "audit-fields", "export",
])
def test_each_cli_help_is_parseable(command, capsys):
    with pytest.raises(SystemExit) as ei:
        main([command, "--help"])
    assert ei.value.code == 0
    assert command in capsys.readouterr().out
