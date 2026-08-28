# -*- coding: utf-8 -*-
import copy
import json
from pathlib import Path

from amazon_es_bestseller.qa.field_closure import audit_field_closure, render_markdown


def _base_product(**overrides):
    p = {
        "asin": "B000000001",
        "title_es_raw": "Caja de almacenamiento",
        "title_zh": "",
        "brand_raw": "",
        "brand": "",
        "current_price_raw": "12,00 €",
        "current_price": 12.0,
        "original_price_raw": "",
        "original_price": None,
        "discount_rate": None,
        "rating_raw": "",
        "rating": None,
        "review_count_raw": "",
        "review_count": None,
        "monthly_bought_raw": "",
        "monthly_bought_min": None,
        "category_l1": "",
        "category_l2": "",
        "category_l3": "",
        "leaf_category": "",
        "bestseller_rank": None,
        "selected_variation_raw": "",
        "spec_v2": "",
        "attributes": [],
        "product_details_zh": "",
        "feature_bullets_raw": [],
        "feature_bullets_zh": "",
        "date_first_available_raw": "",
        "date_first_available": None,
        "seller_raw": "",
        "seller": "",
        "product_url": "https://www.amazon.es/dp/B000000001",
        "image_url": "https://img.example/1.jpg",
        "notes": "",
    }
    p.update(overrides)
    return p


def _issue(report, asin, field):
    return next(r for r in report["records"] if r["asin"] == asin and r["field"] == field)


def test_attribute_brand_without_canonical_is_mapping_missed():
    p = _base_product(attributes=[{"label_raw": "Marca", "value_raw": "De'Longhi"}])
    report = audit_field_closure([p])
    assert _issue(report, p["asin"], "brand")["classification"] == "MAPPING_MISSED"


def test_title_raw_without_chinese_title_is_derived_missing():
    report = audit_field_closure([_base_product()])
    assert _issue(report, "B000000001", "title_zh")["classification"] == "DERIVED_MISSING"


def test_attributes_without_chinese_details_is_derived_missing():
    p = _base_product(attributes=[{"label_raw": "Material", "value_raw": "Acero"}])
    assert _issue(audit_field_closure([p]), p["asin"], "product_details_zh")["classification"] == "DERIVED_MISSING"


def test_bullets_without_chinese_bullets_is_derived_missing():
    p = _base_product(feature_bullets_raw=["Resistente y fácil de limpiar"])
    assert _issue(audit_field_closure([p]), p["asin"], "feature_bullets_zh")["classification"] == "DERIVED_MISSING"


def test_no_source_is_source_missing_and_notes_are_not_audit_field():
    p = _base_product(title_es_raw="", notes="人工备注")
    report = audit_field_closure([p])
    assert _issue(report, p["asin"], "title_zh")["classification"] == "SOURCE_MISSING"
    assert not any(r["field"] in ("notes", "备注") for r in report["records"])


def test_html_evidence_without_raw_is_parser_missed(tmp_path):
    p = _base_product(brand_raw="", brand="")
    (tmp_path / "B000000001.html").write_text(
        '<html><body><div id="bylineInfo">Marca: DeLonghi</div></body></html>',
        encoding="utf-8")
    assert _issue(audit_field_closure([p], html_dir=tmp_path), p["asin"], "brand")["classification"] == "PARSER_MISSED"


def test_page_named_html_is_indexed_by_embedded_asin(tmp_path):
    p = _base_product(brand_raw="", brand="")
    (tmp_path / "page_01.html").write_text(
        '<input id="ASIN" value="B000000001"><div id="bylineInfo">Marca: DeLonghi</div>',
        encoding="utf-8")
    issue = _issue(audit_field_closure([p], html_dir=tmp_path), p["asin"], "brand")
    assert issue["classification"] == "PARSER_MISSED"


def test_multiple_html_directories_are_combined_by_embedded_asin(tmp_path):
    p = _base_product(brand_raw="", brand="")
    first = tmp_path / "home"
    second = tmp_path / "diy"
    first.mkdir()
    second.mkdir()
    (second / "page_01.html").write_text(
        '<input id="ASIN" value="B000000001"><div id="bylineInfo">Marca: DeLonghi</div>',
        encoding="utf-8")
    issue = _issue(audit_field_closure([p], html_dir=[first, second]), p["asin"], "brand")
    assert issue["classification"] == "PARSER_MISSED"


def test_category_without_source_is_source_missing_but_ranking_without_detail_bsr_fallback():
    p = _base_product(detail_bsr_raw="n.º 1 en Hogar y cocina")
    r = {"asin": p["asin"], "bestseller_rank": None, "ranking_source_url": ""}
    report = audit_field_closure([p], rankings=[r])
    assert _issue(report, p["asin"], "category_l1")["classification"] == "SOURCE_MISSING"
    assert _issue(report, p["asin"], "bestseller_rank")["classification"] == "SOURCE_MISSING"


def test_category_source_without_canonical_is_parser_missed():
    p = _base_product()
    r = {"asin": p["asin"], "bestseller_rank": 1,
         "ranking_source_url": "https://www.amazon.es/zgbs/123", "category_path_raw": "Hogar y cocina"}
    assert _issue(audit_field_closure([p], rankings=[r]), p["asin"], "category_l1")["classification"] == "PARSER_MISSED"


def test_run_dir_uses_ranking_html_for_category_evidence(tmp_path):
    p = _base_product()
    (tmp_path / "ranking_00.html").write_text(
        '<div id="zg_browseRoot"><a href="/zgbs/1">Hogar y cocina</a>'
        '<a href="/zgbs/2">Utensilios</a></div>', encoding="utf-8")
    report = audit_field_closure([p], rankings=[{"asin": p["asin"], "bestseller_rank": 1}], run_dir=tmp_path)
    assert _issue(report, p["asin"], "category_l1")["classification"] == "PARSER_MISSED"


def test_category_deeper_level_without_deeper_source_is_source_missing():
    p = _base_product(category_l1="Hogar y cocina")
    report = audit_field_closure([p], rankings=[{"asin": p["asin"], "category_l1": "Hogar y cocina"}])
    assert _issue(report, p["asin"], "category_l2")["classification"] == "SOURCE_MISSING"


def test_generic_variation_script_is_not_selected_variation_evidence(tmp_path):
    p = _base_product()
    (tmp_path / "B000000001.html").write_text(
        '<script>var variationConfig = {"variation":"other"}</script>', encoding="utf-8")
    report = audit_field_closure([p], html_dir=tmp_path)
    assert _issue(report, p["asin"], "selected_variation_raw")["classification"] == "NOT_OBSERVED"


def test_unrelated_struck_variation_price_is_not_original_price_evidence(tmp_path):
    p = _base_product(original_price=None, original_price_raw="")
    (tmp_path / "B000000001.html").write_text(
        '<span class="apex-basisprice-value" data-a-strike="true"><span class="a-offscreen">22,99€</span></span>',
        encoding="utf-8")
    report = audit_field_closure([p], html_dir=tmp_path)
    assert _issue(report, p["asin"], "original_price")["classification"] == "NOT_OBSERVED"


def test_optional_original_price_absent_on_available_page_is_not_observed(tmp_path):
    p = _base_product(original_price_raw="", original_price=None)
    (tmp_path / "B000000001.html").write_text(
        '<input id="ASIN" value="B000000001"><span id="productTitle">Caja</span>',
        encoding="utf-8")
    issue = _issue(audit_field_closure([p], html_dir=tmp_path), p["asin"], "original_price")
    assert issue["classification"] == "NOT_OBSERVED"
    assert issue["severity"] == "INFO"


def test_missing_html_is_evidence_unavailable_not_source_missing():
    issue = _issue(audit_field_closure([_base_product()]), "B000000001", "original_price")
    assert issue["classification"] == "EVIDENCE_UNAVAILABLE"


def test_self_parent_asin_is_not_valid_parent_evidence():
    p = _base_product(parent_asin="")
    report = audit_field_closure([p], details=[{"asin": p["asin"], "parent_asin": p["asin"]}])
    assert _issue(report, p["asin"], "parent_asin")["classification"] == "EVIDENCE_UNAVAILABLE"


def test_self_parent_hidden_input_is_not_page_evidence(tmp_path):
    p = _base_product(parent_asin="")
    (tmp_path / "page_01.html").write_text(
        '<input id="ASIN" value="B000000001"><input id="parentASIN" value="B000000001">',
        encoding="utf-8")
    issue = _issue(audit_field_closure([p], html_dir=tmp_path), p["asin"], "parent_asin")
    assert issue["classification"] == "NOT_OBSERVED"


def test_empty_seller_container_is_not_seller_evidence(tmp_path):
    p = _base_product(seller="", seller_raw="")
    (tmp_path / "page_01.html").write_text(
        '<input id="ASIN" value="B000000001"><div id="productTitle">X</div>'
        '<div id="merchantInfoFeature_feature_div"></div>', encoding="utf-8")
    issue = _issue(audit_field_closure([p], html_dir=tmp_path), p["asin"], "seller")
    assert issue["classification"] == "NOT_OBSERVED"


def test_browse_node_alone_is_not_l3_evidence():
    p = _base_product(category_l3="")
    ranking = {"asin": p["asin"], "category_l1": "Hogar y cocina",
               "category_l2": "Cocina", "category_l3": "", "browse_node_id": "123"}
    issue = _issue(audit_field_closure([p], rankings=[ranking]), p["asin"], "category_l3")
    assert issue["classification"] == "SOURCE_MISSING"


def test_amazon_since_label_is_date_source_evidence(tmp_path):
    p = _base_product(date_first_available_raw="", date_first_available=None)
    (tmp_path / "page_01.html").write_text(
        '<input id="ASIN" value="B000000001">Producto en Amazon.es desde: 6 noviembre 2023',
        encoding="utf-8")
    issue = _issue(audit_field_closure([p], html_dir=tmp_path), p["asin"], "date_first_available")
    assert issue["classification"] == "PARSER_MISSED"


def test_workbook_value_drift_is_export_value_mismatch(tmp_path):
    from openpyxl import load_workbook
    from amazon_es_bestseller.export.excel import export_workbook

    product = _base_product()
    book = tmp_path / "out.xlsx"
    export_workbook([product], out_path=book)
    wb = load_workbook(book)
    wb["西班牙语选品清单"].cell(2, 6).value = 99.0
    wb.save(book)

    report = audit_field_closure([product], workbook_path=book)
    assert any(r["classification"] == "EXPORT_VALUE_MISMATCH" for r in report["records"])


def test_clean_workbook_has_no_export_row_missing_findings(tmp_path):
    from amazon_es_bestseller.export.excel import export_workbook

    product = _base_product(image_url="")
    book = tmp_path / "out.xlsx"
    export_workbook([product], out_path=book)

    report = audit_field_closure([product], workbook_path=book)
    assert not any(r["classification"] == "EXPORT_MISSING" for r in report["records"])


def test_workbook_missing_linked_chinese_image_is_reported(tmp_path):
    from amazon_es_bestseller.export.excel import export_workbook

    product = _base_product()
    book = tmp_path / "out.xlsx"
    export_workbook([product], out_path=book)

    report = audit_field_closure([product], workbook_path=book)
    assert any(r["classification"] == "IMAGE_MISSING" for r in report["records"])


def test_audit_does_not_mutate_records_and_output_is_deterministic():
    p1 = _base_product(asin="B000000002")
    p2 = _base_product(asin="B000000001")
    before = copy.deepcopy([p1, p2])
    one = audit_field_closure([p1, p2])
    two = audit_field_closure([p1, p2])
    assert [r["asin"] for r in one["records"]] == [r["asin"] for r in two["records"]]
    assert one == two
    assert [p1, p2] == before
    assert "Field Closure Audit" in render_markdown(one)


def test_markdown_keeps_coverage_states_out_of_defect_details():
    report = {
        "summary": {},
        "field_summary": {},
        "records": [
            {"asin": "B000000001", "field": "seller", "display_column": "卖家",
             "classification": "NOT_OBSERVED", "severity": "INFO", "raw_evidence": "",
             "canonical_value": "", "display_value": "", "message": "页面未展示"},
            {"asin": "B000000002", "field": "seller", "display_column": "卖家",
             "classification": "PARSER_MISSED", "severity": "P1", "raw_evidence": "",
             "canonical_value": "", "display_value": "", "message": "raw 缺失"},
        ],
    }
    markdown = render_markdown(report)
    assert "页面未展示" not in markdown
    assert "raw 缺失" in markdown


def test_unknown_attribute_evidence_is_preserved_in_audit():
    p = _base_product(attributes=[{"label_raw": "Campo nuevo", "value_raw": "Valor nuevo"}], product_details_zh="Campo nuevo：Valor nuevo")
    report = audit_field_closure([p])
    rec = _issue(report, p["asin"], "product_details_zh")
    assert "Campo nuevo" in json.dumps(rec, ensure_ascii=False)


def test_translation_residual_is_p1():
    p = {"asin": "B000000001", "title_es_raw": "Caja organizadora", "title_zh": "Caja organizadora"}
    report = audit_field_closure([p])
    row = next(r for r in report["records"] if r["field"] == "title_zh")
    assert row["classification"] == "TRANSLATION_INCOMPLETE"


def test_five_sku_golden_fixture_covers_real_issue_shapes():
    fixture = Path(__file__).parent / "fixtures" / "field_closure_golden_5sku.json"
    products = json.loads(fixture.read_text(encoding="utf-8"))["products"]
    report = audit_field_closure(products)
    assert report["summary"]["total_skus"] == 5
    # Detail BSR values are deliberately present, but cannot close bestseller rank.
    no_rank = dict(products[0])
    no_rank["bestseller_rank"] = None
    rank = _issue(audit_field_closure([no_rank]), "B008YETL18", "bestseller_rank")
    assert rank["classification"] == "SOURCE_MISSING"
    invalid = _issue(report, "B07RN64P2R", "original_price")
    assert invalid["classification"] == "ORIGINAL_PRICE_INVALID"
    brand = _issue(report, "B008YETL18", "brand")
    assert brand["classification"] == "MAPPING_MISSED"


def test_basis_price_equal_to_current_is_not_parser_missed(tmp_path):
    """真实回归（文具/宠物 100 SKU 实采）：Amazon 的 "Precio único" 用
    data-a-strike=true 重述当前售价，值与现价完全相同（14,29€ / 14,29€）。

    解析器正确地不把它当划线原价；审计必须复用同一语义，否则会把正确行为
    判成 PARSER_MISSED，并让导出门禁拦下本来合格的数据。实采 100 个 SKU 中
    有 15 个是这种形态。
    """
    from amazon_es_bestseller.qa.field_closure import audit_field_closure
    html = ("<div id='productTitle'>Producto</div><div id='corePrice_feature_div'>"
            "<span class='a-price'><span class='a-offscreen'>14,29€</span></span>"
            "<span class='apex-basisprice-feature'>Precio único: "
            "<span class='a-price a-text-price apex-basisprice-value' data-a-strike='true'>"
            "<span class='a-offscreen'>14,29€</span></span></span></div>")
    record = {"asin": "B000LXUWN6", "current_price": 14.29, "current_price_raw": "14,29€",
              "original_price": None, "original_price_raw": ""}
    detail = {"asin": "B000LXUWN6", "current_price_raw": "14,29€", "original_price_raw": ""}
    html_dir = tmp_path / "html"
    html_dir.mkdir()
    (html_dir / "B000LXUWN6.html").write_text(html, encoding="utf-8")
    report = audit_field_closure([record], details=[detail], rankings=[],
                                 html_dir=[str(html_dir)])
    hits = [r for r in report["records"]
            if r["field"] == "original_price" and r["classification"] == "PARSER_MISSED"]
    assert hits == []
