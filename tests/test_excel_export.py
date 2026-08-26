# -*- coding: utf-8 -*-
"""export/excel.py 测试：B3x 新默认契约（DATA_MODEL §20-§23 / QA_RULES §21-§25）。

断言 3 张表名称/顺序、中文表冻结 26 列、无 配送方式/选品状态/研究备注、
备注按 ASIN 保留、中西表 ASIN 集一致 + 确定性排序、西语表不嵌图、
暂缺字段（完整商品详情/商品卖点/Parent ASIN 等）留空不臆造。
"""
import base64
from io import BytesIO

from openpyxl import load_workbook

from amazon_es_bestseller.export.excel import (
    HEAD_ES,
    HEAD_ZH,
    embed_images_by_asin,
    export_workbook,
    merge_manual_fields,
    to_num,
)

#: 1x1 透明 PNG（测试内嵌图片用）
PNG_1PX = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mP8z8BQDwAEhQGAhKmMIQAAAABJRU5ErkJggg=="
)


def _img3() -> dict:
    """每次调用返回全新 BytesIO 的 3 ASIN 图片表（XLImage 会消费流）。"""
    return {a: (BytesIO(PNG_1PX), 70, 70) for a in
            ('B075JJRFVV', 'B078C6QR1C', 'B07RN64P2R')}


def _zh_header(ws):
    return [ws.cell(row=1, column=c).value for c in range(1, 27)]


def _zh_row(ws, r):
    return [ws.cell(row=r, column=c).value for c in range(1, 27)]


def test_export_3_sheets(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    assert wb.sheetnames == ['类目规划', '西班牙语选品清单', '中文选品清单']


def test_zh_sheet_frozen_26_columns(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    header = _zh_header(ws)
    assert header == HEAD_ZH          # 精确 26 列冻结顺序（DATA_MODEL §21）
    assert len(header) == 26
    for forbidden in ('配送方式', '选品状态', '研究备注'):
        assert not any(h == forbidden for h in header)


def test_zh_column_mapping(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    # ASIN 升序排序后：B075JJRFVV 行2、B078C6QR1C 行3、B07RN64P2R 行4
    row3 = _zh_row(ws, 3)   # B078C6QR1C
    assert row3[1] == 2      # 序号
    assert row3[2] == 'B078C6QR1C'
    assert row3[3] == 'B0DH0ABC01'                       # Parent ASIN
    assert row3[4] == '玻璃便当盒 4 件套'                  # 商品名称（中文）
    assert row3[5] == 'Tatay'                             # 品牌
    assert row3[6] == 12.62                               # 当前售价
    assert row3[7] == 13.29                               # 划线原价
    assert row3[8] == 0.0504                              # 折扣率
    assert row3[9] == 4.5                                 # 评分
    assert row3[10] == 3873                               # 评论数
    assert row3[11] == 500                                # 月购买量
    assert row3[12] == 'Hogar y cocina'                   # 一级类目
    assert row3[13] == 'Almacenamiento y organización'    # 二级类目
    assert row3[14] == 'Juegos de recipientes'            # 三级类目
    assert row3[15] == 'Juegos de recipientes'            # 细分类目
    assert row3[16] == 1                                  # 畅销榜排名
    assert row3[17] == 'Fiambrera - Set 4 Estándar'       # 当前选中规格 / 变体
    assert row3[18] == '4件套'                            # 核心规格（中文）
    assert row3[21] == '2023-10-28'                       # 首次上架日期
    assert row3[22] == 'Tatay'                            # 卖家
    assert row3[23] == 'https://www.amazon.es/dp/B078C6QR1C'
    assert row3[24] == 'https://m.media-amazon.com/images/I/81x.jpg'
    assert row3[25] == '月购看涨（人工）'                   # 备注


def test_zh_detail_cols_empty_without_raw_detail(tmp_path, export_records):
    """无原始全量详情数据（attributes/卖点）→ 列 20/21 留空不臆造（QA_RULES §29）。"""
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    for r in (2, 3, 4):
        assert ws.cell(row=r, column=20).value in (None, '')   # 完整商品详情（中文）
        assert ws.cell(row=r, column=21).value in (None, '')   # 商品卖点（中文）
    # Parent ASIN（第 4 列）缺失（无值）→ 留空，不臆造
    assert ws.cell(row=2, column=4).value in (None, '')        # B075JJRFVV 无 parent
    assert ws.cell(row=4, column=4).value in (None, '')        # B07RN64P2R 无 parent


def test_es_sheet_aligned_and_ordered(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws_es = wb['西班牙语选品清单']
    ws_zh = wb['中文选品清单']
    es_asins = [ws_es.cell(row=r, column=2).value for r in range(2, 5)]
    zh_asins = [ws_zh.cell(row=r, column=3).value for r in range(2, 5)]
    # 中西表 ASIN 集一致、确定性排序（QA_RULES §24）
    assert es_asins == zh_asins == ['B075JJRFVV', 'B078C6QR1C', 'B07RN64P2R']
    # 商品名称（西语）
    assert ws_es.cell(row=3, column=4).value == 'Fiambrera de cristal con 4 piezas'
    # 当前选中规格 / 变体（西语）有值；核心规格西语暂留空；完整详情/卖点西语无原始数据留空
    assert ws_es.cell(row=3, column=17).value == 'Fiambrera - Set 4 Estándar'
    assert ws_es.cell(row=3, column=18).value in (None, '')   # 核心规格（西语）
    assert ws_es.cell(row=3, column=19).value in (None, '')   # 完整商品详情（西语原文）无数据
    assert ws_es.cell(row=3, column=20).value in (None, '')   # 商品卖点（西语原文）无数据


def test_full_detail_rendered_into_cells(tmp_path):
    """带无损全量详情的记录 → 中文列 20/21、西语列 19/20 真实渲染（dedup + 剔除元信息）。"""
    from amazon_es_bestseller.pipeline import normalize_product
    rec = normalize_product({
        "asin": "B008YETL18",
        "attributes": [
            {"section": "product_overview", "label_raw": "Marca", "value_raw": "De'Longhi",
             "position": 0, "source": "productOverview"},
            {"section": "technical_details", "label_raw": "Capacidad", "value_raw": "500 mililitros",
             "position": 0, "source": "prodDetails"},
            # 完全重复 → 渲染层去重
            {"section": "product_overview", "label_raw": "Marca", "value_raw": "De'Longhi",
             "position": 1, "source": "productOverview"},
            # 元信息标签 → 渲染层剔除（原始 attributes 仍在数据层）
            {"section": "technical_details", "label_raw": "ASIN", "value_raw": "B008YETL18",
             "position": 1, "source": "prodDetails"},
        ],
        "feature_bullets_raw": ["Descalcificador para cafeteras", "SOLUCIÓN SUAVE"],
    })
    wb = export_workbook([rec], out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    ws_es = wb['西班牙语选品清单']
    zh_details = ws.cell(row=2, column=20).value      # 完整商品详情（中文）
    zh_bullets = ws.cell(row=2, column=21).value      # 商品卖点（中文）
    assert "品牌：De'Longhi" in zh_details
    assert "容量：500毫升" in zh_details
    assert "ASIN" not in zh_details                   # 元信息已剔除
    assert zh_details.count("品牌：De'Longhi") == 1   # 重复行已去重
    assert "除垢" in zh_bullets                       # 词典关键词翻译
    assert "SOLUCIÓN SUAVE" in zh_bullets             # 未覆盖词保留西语原文
    es_details = ws_es.cell(row=2, column=19).value   # 完整商品详情（西语原文）
    es_bullets = ws_es.cell(row=2, column=20).value   # 商品卖点（西语原文）
    assert "Marca: De'Longhi" in es_details
    assert "Capacidad: 500 mililitros" in es_details
    assert "Descalcificador para cafeteras" in es_bullets


def test_es_sheet_header_25_cols(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['西班牙语选品清单']
    header = [ws.cell(row=1, column=c).value for c in range(1, 26)]
    assert header == HEAD_ES
    for forbidden in ('配送方式', '选品状态', '研究备注'):
        assert forbidden not in header


def test_zh_sheet_images_by_asin(tmp_path, export_records):
    export_workbook(export_records, images_by_asin=_img3(),
                    out_path=str(tmp_path / "out.xlsx"))
    ws = load_workbook(str(tmp_path / "out.xlsx"))['中文选品清单']
    assert len(ws._images) == 3
    # 数据自第 2 行起；ASIN 升序 → 行 2/3/4（._from.row 为 0-based）
    assert sorted(im.anchor._from.row for im in ws._images) == [1, 2, 3]


def test_es_sheet_no_embedded_images(tmp_path, export_records):
    export_workbook(export_records, images_by_asin=_img3(),
                    out_path=str(tmp_path / "out.xlsx"))
    wb = load_workbook(str(tmp_path / "out.xlsx"))
    assert len(wb['西班牙语选品清单']._images) == 0   # 西语表不嵌图（QA_RULES §21）
    assert len(wb['中文选品清单']._images) == 3


def test_embed_images_missing_asin_skipped():
    from openpyxl import Workbook
    ws = Workbook().active
    imgs = {'B0INVALID': (BytesIO(PNG_1PX), 70, 70)}
    embed_images_by_asin(ws, imgs, {'B078C6QR1C': 3})
    assert len(ws._images) == 0


def test_number_formats(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    assert ws.cell(row=3, column=7).number_format == '#,##0.00" €"'   # 当前售价
    assert ws.cell(row=3, column=8).number_format == '#,##0.00" €"'   # 划线原价
    assert ws.cell(row=3, column=9).number_format == '0%'             # 折扣率


def test_hyperlinks_valid(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    for r in (2, 3, 4):
        pcell = ws.cell(row=r, column=24)      # 商品链接
        assert pcell.hyperlink is not None
        assert pcell.hyperlink.target.startswith("https://www.amazon.es/dp/")
        assert pcell.hyperlink.target.endswith(ws.cell(row=r, column=3).value)
        icell = ws.cell(row=r, column=25)      # 图片链接
        assert icell.hyperlink is not None
        assert icell.hyperlink.target.startswith("https://m.media-amazon.com/")


def test_freeze_panes(tmp_path, export_records):
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    assert wb['类目规划'].freeze_panes == 'A2'
    assert wb['西班牙语选品清单'].freeze_panes == 'B2'
    assert wb['中文选品清单'].freeze_panes == 'C2'


def test_merge_manual_fields_preserves_notes(tmp_path, export_records):
    prev = export_workbook(export_records, out_path=str(tmp_path / "prev.xlsx"))

    new_records = [dict(r) for r in export_records]
    for r in new_records:
        if r['asin'] == 'B078C6QR1C':
            r['备注'] = ''          # 新导出未带人工备注
    merged = merge_manual_fields(new_records, prev)
    b078 = next(r for r in merged if r['asin'] == 'B078C6QR1C')
    assert b078['备注'] == '月购看涨（人工）'     # 前版备注按 ASIN 存活


def test_merge_manual_fields_no_clobber(tmp_path, export_records):
    prev = export_workbook(export_records, out_path=str(tmp_path / "prev.xlsx"))

    new_records = [dict(r) for r in export_records]
    for r in new_records:
        if r['asin'] == 'B075JJRFVV':
            r['备注'] = '新人工备注'   # 前版该 ASIN 备注为空 → 不清空不覆盖
    merged = merge_manual_fields(new_records, prev)
    b075 = next(r for r in merged if r['asin'] == 'B075JJRFVV')
    assert b075['备注'] == '新人工备注'


def test_merge_manual_fields_keeps_new_nonempty_note(export_records):
    prev = export_workbook(export_records)
    new_records = [dict(r) for r in export_records]
    new_records[0]["备注"] = "本次人工复核"
    merged = merge_manual_fields(new_records, prev)
    assert merged[0]["备注"] == "本次人工复核"


def test_export_notes_survive_regeneration(tmp_path, export_records):
    """备注经 export_workbook(prev_workbook=...) 端到端按 ASIN 存活（QA_RULES §22）。"""
    prev = export_workbook(export_records, out_path=str(tmp_path / "prev.xlsx"))

    new_records = [dict(r) for r in export_records]
    for r in new_records:
        if r['asin'] == 'B07RN64P2R':
            r['备注'] = ''
    wb = export_workbook(new_records, prev_workbook=prev,
                         out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    # B07RN64P2R → 行 4（ASIN 升序）
    assert ws.cell(row=4, column=26).value == '价格竞争激烈'


def test_translations_fill_title_zh(tmp_path, export_records):
    translations = {
        'B075JJRFVV': {'title_zh': '保温午餐包'},
        'B078C6QR1C': {'title_zh': '翻译表不应覆盖记录已有值'},
    }
    wb = export_workbook(export_records, translations=translations,
                         out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    assert ws.cell(row=2, column=5).value == '保温午餐包'      # 记录无 → 回填翻译表
    assert ws.cell(row=3, column=5).value == '玻璃便当盒 4 件套'  # 记录已有 → 不覆盖


def test_category_planning_sheet(tmp_path, export_records):
    planning = [
        {'#': 1, '中文一级类目': '家居与厨房', 'Amazon 西语名称': 'Hogar y cocina',
         '建议': '主攻', '我的判断': ''},
        {'#': 2, '中文一级类目': '玩具', 'Amazon 西语名称': 'Juguetes',
         '建议': '观察', '我的判断': '竞品多'},
    ]
    wb = export_workbook(export_records, category_planning=planning,
                         out_path=str(tmp_path / "out.xlsx"))
    ws = wb['类目规划']
    header = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    assert header == ['#', '中文一级类目', 'Amazon 西语名称', '建议', '我的判断']
    assert ws.cell(row=2, column=2).value == '家居与厨房'
    assert ws.cell(row=3, column=5).value == '竞品多'


def test_export_deterministic_order_by_asin(tmp_path, export_records):
    # export_records 输入故意未排序（B078C6QR1C 在前）；导出必须 ASIN 升序
    wb = export_workbook(export_records, out_path=str(tmp_path / "out.xlsx"))
    ws = wb['中文选品清单']
    assert [ws.cell(row=r, column=3).value for r in range(2, 5)] == \
        ['B075JJRFVV', 'B078C6QR1C', 'B07RN64P2R']
    ws_es = wb['西班牙语选品清单']
    assert [ws_es.cell(row=r, column=2).value for r in range(2, 5)] == \
        ['B075JJRFVV', 'B078C6QR1C', 'B07RN64P2R']


def test_export_does_not_mutate_input(export_records):
    snapshot = [dict(r) for r in export_records]
    export_workbook(export_records)
    assert export_records == snapshot


def test_to_num():
    assert to_num('12,62 €') == 12.62
    assert to_num('13,29') == 13.29
    assert to_num(0.0504) == 0.0504
    assert to_num('') is None
    assert to_num(None) is None
    assert to_num('abc') is None
