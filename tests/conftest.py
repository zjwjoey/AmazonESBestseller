# -*- coding: utf-8 -*-
"""pytest 共享 fixtures。所有 fixture 均为离线数据，测试绝不联网。"""
from pathlib import Path

import openpyxl
import pytest

FIXTURES_DIR = Path(__file__).parent / "fixtures"
HTML_DIR = FIXTURES_DIR / "html"


def read_html(name: str) -> str:
    return (HTML_DIR / name).read_text(encoding="utf-8")


@pytest.fixture
def lunchbag_html() -> str:
    return read_html("product_lunchbag.html")


@pytest.fixture
def captcha_html() -> str:
    return read_html("product_captcha.html")


@pytest.fixture
def bestsellers_grid_html() -> str:
    return read_html("bestsellers_grid.html")


@pytest.fixture
def sample_bsr_text() -> str:
    return (
        "nº52 en Hogar y cocina ( Ver el Top 100 en Hogar y cocina ) "
        "nº1 en Juegos de recipientes"
    )


@pytest.fixture
def sample_detail_dict() -> dict:
    """与 prep_v2_selection.py 处理的 details_json 结构一致的真实字段子集。"""
    return {
        "capacidad": "30 l",
        "dimensiones_del_articulo_largo_x_ancho_x_alto": "10 l. x 15 an. x 5 al. centímetros",
        "numero_de_articulos": "4",
        "numero_de_sets": "4",
        "tipo_de_material": "Acero inoxidable",
        "clasificacion_en_los_mas_vendidos_de_amazon": (
            "nº52 en Hogar y cocina ( Ver el Top 100 en Hogar y cocina ) "
            "nº1 en Juegos de recipientes"
        ),
    }


@pytest.fixture
def tiny_records() -> list[dict]:
    """3 行商品记录（含中文派生字段与人工字段），供导出测试使用。"""
    return [
        {
            "asin": "B078C6QR1C",
            "采集类目中文": "家居与厨房",
            "title_es_raw": "Fiambrera de cristal con 4 piezas",
            "brand": "Tatay",
            "current_price": "12,62",
            "original_price": "13,29",
            "discount_rate": "0.0504",
            "rating": "4,5",
            "review_count": "3873",
            "monthly_bought_min": "",
            "bsr_leaf_cat": "Juegos de recipientes",
            "bsr_leaf_rank": "1",
            "spec_v2": "4件套",
            "summary_v2": "材质：不锈钢",
            "date_first_available": "2023-10-28",
            "product_url": "https://www.amazon.es/dp/B078C6QR1C",
            "image_url": "https://m.media-amazon.com/images/I/81x.jpg",
            "选品状态": "重点关注",
            "研究备注": "月购看涨",
        },
        {
            "asin": "B075JJRFVV",
            "采集类目中文": "家居与厨房",
            "title_es_raw": "Bolsa térmica para comer",
            "brand": "Utopia Bedding",
            "current_price": "16,98",
            "original_price": "",
            "discount_rate": "",
            "rating": "4,2",
            "review_count": "12455",
            "monthly_bought_min": "",
            "bsr_leaf_cat": "Fiambreras y bolsas de almuerzo",
            "bsr_leaf_rank": "2",
            "spec_v2": "",
            "summary_v2": "材质：涤纶",
            "date_first_available": "2021-03-14",
            "product_url": "https://www.amazon.es/dp/B075JJRFVV",
            "image_url": "https://m.media-amazon.com/images/I/82y.jpg",
            "选品状态": "待评估",
            "研究备注": "",
        },
        {
            "asin": "B07RN64P2R",
            "采集类目中文": "家居与厨房",
            "title_es_raw": "Juego de recipientes reutilizables",
            "brand": "Amazon Basics",
            "current_price": "13,52",
            "original_price": "15,00",
            "discount_rate": "0.0987",
            "rating": "4,6",
            "review_count": "2143",
            "monthly_bought_min": "100",
            "bsr_leaf_cat": "Juegos de recipientes",
            "bsr_leaf_rank": "3",
            "spec_v2": "6件套 / 500毫升",
            "summary_v2": "材质：玻璃",
            "date_first_available": "2020-01-09",
            "product_url": "https://www.amazon.es/dp/B07RN64P2R",
            "image_url": "https://m.media-amazon.com/images/I/83z.jpg",
            "选品状态": "已研究",
            "研究备注": "价格竞争激烈",
        },
    ]


@pytest.fixture
def export_records() -> list[dict]:
    """3 行商品记录（新导出契约 shape，B3x），供导出测试使用。

    与 tiny_records（QA 测试用）独立：本 fixture 带新契约 canonical 字段
    （category_l1..l3 / leaf_category / title_zh / parent_asin / seller_raw /
    bestseller_rank / selected_variation_raw / 备注）。输入故意**未按 ASIN 排序**
    （B078C6QR1C 在前），导出内部做确定性排序（QA_RULES §24）。
    """
    return [
        {
            "asin": "B078C6QR1C",
            "parent_asin": "B0DH0ABC01",
            "title_es_raw": "Fiambrera de cristal con 4 piezas",
            "title_zh": "玻璃便当盒 4 件套",
            "brand": "Tatay",
            "current_price": 12.62,
            "original_price": 13.29,
            "discount_rate": 0.0504,
            "rating": 4.5,
            "review_count": 3873,
            "monthly_bought_min": 500,
            "category_l1": "Hogar y cocina",
            "category_l2": "Almacenamiento y organización",
            "category_l3": "Juegos de recipientes",
            "leaf_category": "Juegos de recipientes",
            "bestseller_rank": 1,
            "selected_variation_raw": "Fiambrera - Set 4 Estándar",
            "spec_v2": "4件套",
            "date_first_available": "2023-10-28",
            "seller_raw": "Tatay",
            "product_url": "https://www.amazon.es/dp/B078C6QR1C",
            "image_url": "https://m.media-amazon.com/images/I/81x.jpg",
            "备注": "月购看涨（人工）",
        },
        {
            "asin": "B075JJRFVV",
            "parent_asin": "",
            "title_es_raw": "Bolsa térmica para comer",
            "title_zh": "",
            "brand": "Utopia Bedding",
            "current_price": 16.98,
            "original_price": "",
            "discount_rate": None,
            "rating": 4.2,
            "review_count": 12455,
            "monthly_bought_min": "",
            "category_l1": "Hogar y cocina",
            "category_l2": "Almacenamiento y organización",
            "category_l3": "Bolsas de almuerzo",
            "leaf_category": "Bolsas de almuerzo",
            "bestseller_rank": 2,
            "selected_variation_raw": "",
            "spec_v2": "",
            "date_first_available": "2021-03-14",
            "seller_raw": "Utopia Brands",
            "product_url": "https://www.amazon.es/dp/B075JJRFVV",
            "image_url": "https://m.media-amazon.com/images/I/82y.jpg",
            "备注": "",
        },
        {
            "asin": "B07RN64P2R",
            "parent_asin": None,
            "title_es_raw": "Juego de recipientes reutilizables",
            "title_zh": "可重复使用收纳盒套装",
            "brand": "Amazon Basics",
            "current_price": 13.52,
            "original_price": 15.00,
            "discount_rate": 0.0987,
            "rating": 4.6,
            "review_count": 2143,
            "monthly_bought_min": 100,
            "category_l1": "Hogar y cocina",
            "category_l2": "Almacenamiento y organización",
            "category_l3": "Juegos de recipientes",
            "leaf_category": "Organizadores de despensa",
            "bestseller_rank": 3,
            "selected_variation_raw": "Juego de 6",
            "spec_v2": "6件套 / 500毫升",
            "date_first_available": "2020-01-09",
            "seller_raw": "Amazon",
            "product_url": "https://www.amazon.es/dp/B07RN64P2R",
            "image_url": "https://m.media-amazon.com/images/I/83z.jpg",
            "备注": "价格竞争激烈",
        },
    ]


def read_xlsx(path: str | Path):
    """读取导出的 xlsx 返回 (workbook, sheetnames)，供导出测试检查。"""
    wb = openpyxl.load_workbook(path, data_only=True)
    return wb, wb.sheetnames
